"""ローカル .xlsx への書き戻し（openpyxl）。privileged。

既定は in-place 上書き。`output` 指定時のみ別名保存。
"""

from __future__ import annotations

from typing import Optional

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple


def _anchor(ws, cell: str) -> str:
    """cell が結合範囲内なら左上アンカーに解決する（結合セルは左上のみ書込可）。"""
    row, col = coordinate_to_tuple(cell)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell


def set_cell_text(
    path: str, sheet: str, cell: str, text: str, *, output: Optional[str] = None
) -> str:
    wb = load_workbook(path)
    ws = wb[sheet]
    ws[_anchor(ws, cell)] = text
    dest = output or path
    wb.save(dest)
    return dest


def set_cell_image(
    path: str,
    sheet: str,
    anchor: str,
    image_path: str,
    *,
    width: Optional[int] = None,
    height: Optional[int] = None,
    output: Optional[str] = None,
) -> str:
    wb = load_workbook(path)
    ws = wb[sheet]
    img = XLImage(image_path)
    if width:
        img.width = width
    if height:
        img.height = height
    img.anchor = anchor  # 左上アンカー（例 "B2"）
    ws.add_image(img)
    dest = output or path
    wb.save(dest)
    return dest


def write_cells(
    path: str, sheet: str, assignments: list[tuple[str, object]], *, output: Optional[str] = None
) -> str:
    """[(セル, 値), ...] を一括書込（構造書き戻し用）。"""
    wb = load_workbook(path)
    ws = wb[sheet]
    for cell, value in assignments:
        ws[_anchor(ws, cell)] = value
    dest = output or path
    wb.save(dest)
    return dest
