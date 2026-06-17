"""testspec 追記ブロックを既存テンプレの体裁に整形する（罫線/縦結合/番号）。privileged。

gwex の書き込みは値だけを置く。実テンプレは「全セル細罫線の箱＋親（画面/中項目/小項目）
を縦結合＋No 列」という体裁。本モジュールは追記済みブロックに対し、その体裁を後付けする。
全セルに箱罫線を引き、親領域を結合すると、結合領域は外枠のみ・ケースセルは個別箱になり
既存行と同じ「区切り線」になる。フォント/配置は直上の既存行から複製して揃える。
"""

from __future__ import annotations

import copy
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from gwex.domains.model import DEFAULT_MAPPING, MappingConfig


def _box_cols(cfg: MappingConfig) -> tuple[int, int]:
    """箱（罫線）の左右端の列インデックス(1始まり)を返す。"""
    if cfg.box_left_col:
        left = column_index_from_string(cfg.box_left_col)
    else:
        cand = [cfg.column_index(r) for r in ("screen_name",)] + [
            column_index_from_string(c) - 1 for c in (cfg.number_columns or {}).values()
        ]
        left = min(c for c in cand if c is not None) + 1
    if cfg.box_right_col:
        right = column_index_from_string(cfg.box_right_col)
    else:
        vals = [cfg.column_index(r) for r in (
            "test_no", "verification_content", "execution_steps")]
        right = max(c for c in vals if c is not None) + 1
    return left, right


def format_testspec_block(
    path: str,
    sheet: str,
    layout: dict,
    mapping: Optional[MappingConfig] = None,
    *,
    output: Optional[str] = None,
) -> str:
    """追記ブロック（layout）に罫線・縦結合・親番号を適用する。

    layout は domains.testspec_writer.layout() の戻り値。
    """
    cfg = mapping or DEFAULT_MAPPING
    wb = load_workbook(path)
    ws = wb[sheet]
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    r0, r1 = layout["data_first_row"], layout["data_last_row"]
    left, right = _box_cols(cfg)

    # スタイル複製元の選定: 直上の既存行を使うが、それがヘッダ（塗りつぶしあり）の場合は
    # データ行の体裁を引きずってしまう（M8/M11: データセルは塗りなし・left/top/wrap が正）。
    # 3段ヘッダ直後にデータが始まるテンプレでは r0-1 がヘッダ行になるため、
    # ヘッダを掴んだら複製せず、テンプレ標準のデータセル体裁を明示適用する。
    from openpyxl.styles import Alignment, PatternFill

    ref_row = r0 - 1
    ref_is_header = any(
        ws.cell(row=ref_row, column=c).fill.patternType for c in range(left, right + 1)
    )
    no_fill = PatternFill(fill_type=None)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for r in range(r0, r1 + 1):
        for c in range(left, right + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = box
            ref = ws.cell(row=ref_row, column=c)
            if ref_is_header:
                # データセル標準体裁（塗りなし・太字なし・left/top/wrap）を明示適用
                if ref.has_style:
                    f = copy.copy(ref.font)
                    f.bold = False
                    cell.font = f
                    cell.number_format = ref.number_format
                cell.alignment = copy.copy(data_align)
                cell.fill = copy.copy(no_fill)
            elif ref.has_style:
                cell.font = copy.copy(ref.font)
                cell.alignment = copy.copy(ref.alignment)
                cell.number_format = ref.number_format
                cell.fill = copy.copy(ref.fill)

    for cell, no in layout["numbers"]:
        ws[cell] = no

    for ref in layout["merges"]:
        ws.merge_cells(ref)

    # データ行の customHeight を解除 → Excel がファイルを開くとき wrap_text に合わせて自動フィット
    for r in range(r0, r1 + 1):
        ws.row_dimensions[r].height = None  # customHeight=False になる

    dest = output or path
    wb.save(dest)
    return dest


_DEFAULT_ROW_PT = 15.75  # 既定行高(pt)


def _col_px(ws, c: int) -> int:
    w = ws.column_dimensions[get_column_letter(c)].width
    return int((w or 8.43) * 7) + 5


def create_before_after_section(
    path: str,
    sheet: str,
    top_row: int,
    title: str,
    before_image: Optional[str] = None,
    after_image: Optional[str] = None,
    *,
    left_cols: tuple[str, str] = ("C", "L"),     # セクション全体の左右端
    split_col: str = "H",                          # 修正後の開始列（修正前=left..split-1）
    row_height_pt: float = _DEFAULT_ROW_PT,
    header_fill: str = "FF9CC2E5",                  # 見出し塗り（既存テンプレ色, ARGB）
    area_fill: str = "FFDEEAF6",                    # ラベル/画像エリア塗り（薄い青, ARGB）
    n_rows: Optional[int] = None,                   # 画像エリアの行数を固定（Noneの場合は画像比率から自動計算）
    output: Optional[str] = None,
) -> str:
    """既存テンプレ準拠の「見出し＋修正前/修正後＋箱罫線」セクション。

    画像は **横幅=各カラム群の幅で固定**し、**縦は画像の自然高さに収まるだけ行を確保**する
    （空行で間延びさせない）。n_rows を指定すると行数を固定できる（複数ブロックで統一したい場合）。
    before_image/after_image 両方 None の場合は枠のみ作成（画像なし）。
    罫線/結合/ラベルは openpyxl、画像は xlsx_zip で最後に oneCellAnchor
    配置（ext を確定させ、以後の openpyxl 保存でのサイズリセットを避ける）。
    """
    from gwex.writer import xlsx_zip

    lc = column_index_from_string(left_cols[0])
    rc = column_index_from_string(left_cols[1])
    sc = column_index_from_string(split_col)
    r_label = top_row + 1
    r_img = top_row + 2

    from openpyxl.styles import PatternFill

    wb = load_workbook(path)
    ws = wb[sheet]

    # 既存テンプレ準拠の塗り（見出し=濃い青、ラベル/画像エリア=薄い青）
    fill_head = PatternFill(fgColor=header_fill, fill_type="solid")
    fill_area = PatternFill(fgColor=area_fill, fill_type="solid")

    w_before = sum(_col_px(ws, c) for c in range(lc, sc))
    w_after = sum(_col_px(ws, c) for c in range(sc, rc + 1))
    box_w = min(w_before, w_after)
    row_px = max(1, int(round(row_height_pt * 96 / 72)))

    ref_img = before_image or after_image
    if ref_img is not None:
        from PIL import Image as PILImage
        iw, ih = PILImage.open(ref_img).size
        box_h = int(box_w * ih / iw)                      # 箱の高さ（画像比率）
        n_rows_val = n_rows if n_rows is not None else max(1, round(box_h / row_px))
        per_row_pt = (box_h / n_rows_val) * 72 / 96       # 合計 = box_h px
    else:
        # 画像なし: 枠のみ作成。n_rows未指定なら34行デフォルト
        n_rows_val = n_rows if n_rows is not None else 34
        per_row_pt = row_height_pt
        box_h = int(n_rows_val * row_px)

    r_bottom = r_img + n_rows_val - 1
    for r in range(r_img, r_bottom + 1):
        ws.row_dimensions[r].height = per_row_pt

    from openpyxl.styles import Alignment
    center = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # このセクションが占める行列範囲に重複する既存結合セルを先に解除する
    from openpyxl.utils import get_column_letter
    rows_span = range(top_row, r_bottom + 1)
    cols_span = range(lc, rc + 1)
    to_unmerge = [
        str(m) for m in list(ws.merged_cells.ranges)
        if any(m.min_row <= r <= m.max_row for r in rows_span)
        and any(m.min_col <= c <= m.max_col for c in cols_span)
    ]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)

    ws.merge_cells(start_row=top_row, start_column=lc, end_row=top_row, end_column=rc)
    ws.cell(row=top_row, column=lc).value = title
    ws.cell(row=top_row, column=lc).alignment = center
    ws.merge_cells(start_row=r_label, start_column=lc, end_row=r_label, end_column=sc - 1)
    ws.cell(row=r_label, column=lc).value = "修正前"
    ws.cell(row=r_label, column=lc).alignment = center
    ws.merge_cells(start_row=r_label, start_column=sc, end_row=r_label, end_column=rc)
    ws.cell(row=r_label, column=sc).value = "修正後"
    ws.cell(row=r_label, column=sc).alignment = center
    for r in range(top_row, r_bottom + 1):
        for c in range(lc, rc + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = box
            cell.fill = fill_head if r == top_row else fill_area
    # 画像エリアのセルを結合（内部格子を消して 1 枚の青パネルに）: 修正前 / 修正後
    ws.merge_cells(start_row=r_img, start_column=lc, end_row=r_bottom, end_column=sc - 1)
    ws.merge_cells(start_row=r_img, start_column=sc, end_row=r_bottom, end_column=rc)
    dest = output or path
    wb.save(dest)

    # 画像は枠の 80% サイズ・比率維持で中央に置く（青背景の上）。両画像とも同寸。
    disp_w = max(1, int(box_w * 0.8))
    disp_h = max(1, int(box_h * 0.8))            # = disp_w * ih/iw（比率維持）

    def place(col_letter, img_path):
        c0 = column_index_from_string(col_letter)
        c1 = (sc - 1) if col_letter == left_cols[0] else rc
        sub_w = sum(_col_px(ws, c) for c in range(c0, c1 + 1))
        col_off = int((sub_w - disp_w) / 2)      # サブ枠内で水平中央
        row_off = int((box_h - disp_h) / 2)      # 画像エリア内で垂直中央
        xlsx_zip.set_cell_image(
            dest, sheet, f"{col_letter}{r_img}", img_path,
            width=disp_w, height=disp_h, col_off_px=col_off, row_off_px=row_off,
            max_dim=None, output=dest,
        )

    if before_image:
        place(left_cols[0], before_image)
    if after_image:
        place(split_col, after_image)
    return dest
