"""xlsx を ZIP として扱い、既存内容に最小タッチで over-cell 画像を注入する（privileged）。

なぜ openpyxl(load+save) でなくこれを使うか：openpyxl の round-trip は**任意の実 Excel** で
グラフ・ピボット・スライサー・フォームコントロール・VBA(xlsm) 等の高度要素をドロップ/破損
させる既知の弱点があり、また既存 media を再圧縮・再採番し sharedStrings/calcChain/xl/metadata
を書き換える。本モジュールは対象シート＋drawing 以外の ZIP エントリに一切触れず、既存の
media/書式/高度要素を**バイト単位で温存**する（対象が単純な doc なら openpyxl でも実害は無いが、
任意 doc への安全側の既定として ZIP 注入を採る）。

※「openpyxl は埋め込み画像を破壊する」という旧前提は誤り。現行 openpyxl は画像(ICC含む)を
保持する（破壊はしないが上記の書き換え/高度要素ロスは起きる）。

`set_cell_image` は xlsx_writer.set_cell_image と同じシグネチャ・同じ振る舞い
（cell_range 枠フィット＝アスペクト維持、width/height 明示、重複結合の解除、
insert_rows で枠ぶんの行挿入）を、既存画像を壊さずに実現する。
"""

from __future__ import annotations

import io
import re
import zipfile
from typing import Optional

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from gwex.writer._image import downscale

_EMU = 9525  # 1px
_COL_PX = 64.0   # 既定列幅相当(px)。xlsx_writer と一致させる
_ROW_PX = 20.0   # 既定行高相当(px)


def _read_zip(path: str) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as z:
        return {n: z.read(n) for n in z.namelist()}


def _write_zip(entries: dict[str, bytes], dest: str) -> None:
    with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in entries.items():
            z.writestr(name, data)


def _sheet_xml_path(entries: dict[str, bytes], sheet_name: str) -> str:
    wb = entries["xl/workbook.xml"].decode("utf-8")
    m = re.search(r'<sheet[^>]*name="' + re.escape(sheet_name) + r'"[^>]*r:id="(rId\d+)"', wb)
    if not m:
        raise ValueError(f"シートが見つかりません: {sheet_name}")
    rid = m.group(1)
    rels = entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
    # 属性順は実装依存（Target が Id より前のこともある）ので、当該 Relationship 要素を
    # 取り出してから Target を読む。
    rel = re.search(r'<Relationship\b[^>]*\bId="' + re.escape(rid) + r'"[^>]*/>', rels)
    if not rel:
        raise ValueError(f"worksheet rel が見つかりません: {rid}")
    mt = re.search(r'Target="([^"]+)"', rel.group(0))
    if not mt:
        raise ValueError(f"worksheet rel に Target がありません: {rid}")
    target = mt.group(1).lstrip("/")
    return target if target.startswith("xl/") else "xl/" + target


def _next_media_index(entries: dict[str, bytes]) -> int:
    idx = 0
    for n in entries:
        m = re.match(r"xl/media/image(\d+)\.", n)
        if m:
            idx = max(idx, int(m.group(1)))
    return idx + 1


def _ensure_png_content_type(entries: dict[str, bytes]) -> None:
    ct = entries["[Content_Types].xml"].decode("utf-8")
    if 'Extension="png"' not in ct:
        ct = ct.replace("</Types>", '<Default Extension="png" ContentType="image/png"/></Types>')
        entries["[Content_Types].xml"] = ct.encode("utf-8")


def _find_drawing(entries: dict[str, bytes], sheet_xml: str) -> Optional[tuple[str, str]]:
    """シートに配線済みの drawing パスと drawing rels パスを返す（無ければ None）。"""
    base = sheet_xml.split("/")[-1]  # sheet2.xml
    rels_path = f"xl/worksheets/_rels/{base}.rels"
    rels = entries.get(rels_path, b"").decode("utf-8")
    m = re.search(r'Target="([^"]*drawings/drawing\d+\.xml)"', rels) if rels else None
    if not m:
        return None
    draw = m.group(1).split("/")[-1]
    return f"xl/drawings/{draw}", f"xl/drawings/_rels/{draw}.rels"


def _drawing_for_sheet(entries: dict[str, bytes], sheet_xml: str) -> tuple[str, str]:
    """シートの drawing パスと drawing rels パスを返す。無ければ新規作成して配線する。"""
    found = _find_drawing(entries, sheet_xml)
    if found:
        return found

    # 新規 drawing を作成
    rels_path = f"xl/worksheets/_rels/{sheet_xml.split('/')[-1]}.rels"
    nums = [int(x) for n in entries for x in re.findall(r"xl/drawings/drawing(\d+)\.xml$", n)]
    dn = (max(nums) + 1) if nums else 1
    draw_path = f"xl/drawings/drawing{dn}.xml"
    draw_rels = f"xl/drawings/_rels/drawing{dn}.xml.rels"
    entries[draw_path] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"></xdr:wsDr>'
    ).encode("utf-8")
    entries[draw_rels] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>'
    ).encode("utf-8")
    # シートに drawing rel と <drawing> 要素を追加
    rid = _add_rel(entries, rels_path,
                   "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
                   f"../drawings/drawing{dn}.xml")
    sx = entries[sheet_xml].decode("utf-8")
    if "<drawing " not in sx:
        # <worksheet> に r: 名前空間が無い（新規 openpyxl シート等）と <drawing r:id=>
        # の r: が未束縛になりパース不能。無ければ宣言を補う。
        wm = re.match(r"(<worksheet\b[^>]*?)>", sx)
        if wm and "xmlns:r=" not in wm.group(1):
            sx = sx.replace(
                wm.group(0),
                wm.group(1)
                + ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
                1,
            )
        sx = re.sub(r"</worksheet>", f'<drawing r:id="{rid}"/></worksheet>', sx)
        entries[sheet_xml] = sx.encode("utf-8")
    # [Content_Types] に drawing を宣言
    ct = entries["[Content_Types].xml"].decode("utf-8")
    if f"/xl/drawings/drawing{dn}.xml" not in ct:
        ct = ct.replace("</Types>",
                        f'<Override PartName="/xl/drawings/drawing{dn}.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
        entries["[Content_Types].xml"] = ct.encode("utf-8")
    return draw_path, draw_rels


def _expand_self_closed_wsdr(dx: str) -> str:
    """空の drawing が `<xdr:wsDr .../>`（自己終了）で書かれていると、`</xdr:wsDr>` を
    探す追記ロジックが黙って何もしない。開始タグ＋閉じタグの形に正規化する。

    テンプレ（Excel が書いた xlsx）の未使用シートの drawing は実際に自己終了形式。
    """
    m = re.match(r"(\s*<\?xml[^>]*\?>)?\s*<((?:\w+:)?wsDr)\b([^>]*?)/>\s*$", dx, re.S)
    if not m:
        return dx
    decl = m.group(1) or ""
    tag, attrs = m.group(2), m.group(3)
    return f"{decl}<{tag}{attrs}></{tag}>"


_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"


def _ensure_drawing_namespaces(dx: str) -> str:
    """openpyxl が書いた drawing は既定名前空間（接頭辞なし）のことがある。
    xdr:/a:/r: 接頭辞のアンカーを差し込む前に、ルートへ宣言を補う。"""
    m = re.search(r"<(?:xdr:)?wsDr\b[^>]*>", dx)
    if not m:
        return dx
    tag = m.group(0)
    add = ""
    if "xmlns:xdr=" not in tag:
        add += f' xmlns:xdr="{_XDR_NS}"'
    if "xmlns:a=" not in tag:
        add += f' xmlns:a="{_A_NS}"'
    if "xmlns:r=" not in tag:
        add += f' xmlns:r="{_R_NS}"'
    if not add:
        return dx
    new_tag = tag[:-1].rstrip() + add + ">"
    return dx[: m.start()] + new_tag + dx[m.end():]


_ANCHOR_RE = re.compile(
    r"<(?:xdr:)?(oneCellAnchor|twoCellAnchor|absoluteAnchor)\b.*?</(?:xdr:)?\1>", re.S)


def _shape_anchors(dx: str) -> list[str]:
    """drawing XML から「図形（sp/grpSp）を含むアンカー」だけを抜き出す。画像アンカーは対象外。"""
    return [m.group(0) for m in _ANCHOR_RE.finditer(dx)
            if "<xdr:sp" in m.group(0) or "<xdr:grpSp" in m.group(0)]


def _max_shape_id(dx: str) -> int:
    ids = [int(x) for x in re.findall(r'<xdr:cNvPr[^>]*\bid="(\d+)"', dx)]
    return max(ids) if ids else 0


def _renumber_ids(xml: str, start: int) -> str:
    """cNvPr の id を start から振り直す（同一 drawing 内で id は一意である必要がある）。"""
    counter = [start]

    def sub(m):
        counter[0] += 1
        return m.group(1) + str(counter[0]) + m.group(3)

    return re.sub(r'(<xdr:cNvPr[^>]*\bid=")(\d+)(")', sub, xml)


def _add_rel(entries: dict[str, bytes], rels_path: str, rtype: str, target: str) -> str:
    rels = entries.get(rels_path, b"").decode("utf-8")
    if not rels:
        rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>')
    existing = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels)]
    rid = f"rId{(max(existing) + 1) if existing else 1}"
    rel = f'<Relationship Id="{rid}" Type="{rtype}" Target="{target}"/>'
    rels = rels.replace("</Relationships>", rel + "</Relationships>")
    entries[rels_path] = rels.encode("utf-8")
    return rid


def _add_merge(entries: dict[str, bytes], sheet_xml: str, ref: str) -> None:
    sx = entries[sheet_xml].decode("utf-8")
    m = re.search(r'<mergeCells count="(\d+)">', sx)
    if m:
        cnt = int(m.group(1)) + 1
        sx = sx.replace(m.group(0), f'<mergeCells count="{cnt}">', 1)
        sx = sx.replace("</mergeCells>", f'<mergeCell ref="{ref}"/></mergeCells>', 1)
    else:
        # sheetData の後ろに mergeCells を新設
        sx = re.sub(r"</sheetData>",
                    f'</sheetData><mergeCells count="1"><mergeCell ref="{ref}"/></mergeCells>', sx, count=1)
    entries[sheet_xml] = sx.encode("utf-8")


def _remove_overlapping_merges(entries: dict[str, bytes], sheet_xml: str,
                               min_c: int, min_r: int, max_c: int, max_r: int) -> None:
    """指定範囲と重なる既存 mergeCell を取り除く（重複結合の解除）。"""
    sx = entries[sheet_xml].decode("utf-8")
    removed = 0

    def keep(m):
        nonlocal removed
        a, b = m.group(1), m.group(2)
        c1, r1 = coordinate_to_tuple(a)[1], coordinate_to_tuple(a)[0]
        if b:
            c2, r2 = coordinate_to_tuple(b)[1], coordinate_to_tuple(b)[0]
        else:
            c2, r2 = c1, r1
        lo_c, hi_c, lo_r, hi_r = min(c1, c2), max(c1, c2), min(r1, r2), max(r1, r2)
        overlap = not (hi_c < min_c or lo_c > max_c or hi_r < min_r or lo_r > max_r)
        if overlap:
            removed += 1
            return ""
        return m.group(0)

    sx = re.sub(r'<mergeCell ref="([A-Z]+\d+)(?::([A-Z]+\d+))?"/>', keep, sx)
    if removed:
        mc = re.search(r'<mergeCells count="(\d+)">', sx)
        if mc:
            new_cnt = int(mc.group(1)) - removed
            if new_cnt <= 0:
                sx = re.sub(r'<mergeCells count="\d+">\s*</mergeCells>', "", sx)
            else:
                sx = sx.replace(mc.group(0), f'<mergeCells count="{new_cnt}">', 1)
    entries[sheet_xml] = sx.encode("utf-8")


def _shift_rows_down(entries: dict[str, bytes], sheet_xml: str, draw_path: str,
                     min_r: int, nrows: int) -> None:
    """min_r 以降の行を nrows 行ぶん下へずらす（既存内容を押し下げ、空枠を作る）。

    sheetData の <row r=> と セル <c r=>、mergeCell の参照、既存 drawing アンカーの
    <xdr:row>（0始まり）をまとめてシフトする。
    """
    sx = entries[sheet_xml].decode("utf-8")

    # 1) 行タグ <row ... r="N" ...>
    def shift_rowtag(m):
        n = int(m.group(2))
        return f'{m.group(1)}{n + nrows if n >= min_r else n}{m.group(3)}'
    sx = re.sub(r'(<row\b[^>]*?\br=")(\d+)(")', shift_rowtag, sx)

    # 2) セル参照 <c r="B7" ...>（列文字＋行番号）
    def shift_cellref(m):
        col, n = m.group(1), int(m.group(2))
        return f'r="{col}{n + nrows if n >= min_r else n}"'
    sx = re.sub(r'\br="([A-Z]+)(\d+)"', shift_cellref, sx)

    # 3) mergeCell ref="A5:C8"
    def shift_one(ref):
        cm = re.match(r'([A-Z]+)(\d+)', ref)
        col, n = cm.group(1), int(cm.group(2))
        return f'{col}{n + nrows if n >= min_r else n}'

    def shift_merge(m):
        parts = m.group(1).split(":")
        return '<mergeCell ref="' + ":".join(shift_one(p) for p in parts) + '"/>'
    sx = re.sub(r'<mergeCell ref="([A-Z0-9:]+)"/>', shift_merge, sx)
    entries[sheet_xml] = sx.encode("utf-8")

    # 4) 既存 drawing アンカー <xdr:row>N</xdr:row>（0始まり）
    if draw_path in entries:
        dx = entries[draw_path].decode("utf-8")
        thr = min_r - 1  # 0始まりしきい値

        def shift_anchor(m):
            pre, n = m.group(1), int(m.group(2))
            return f'<{pre}row>{n + nrows if n >= thr else n}</{pre}row>'
        # xdr: 接頭辞（Excel）と接頭辞なし（openpyxl）の両対応
        dx = re.sub(r'<(xdr:)?row>(\d+)</(?:xdr:)?row>', shift_anchor, dx)
        entries[draw_path] = dx.encode("utf-8")


def _sheet_defaults(sx: str) -> tuple[float, float]:
    """defaultColWidth(文字幅), defaultRowHeight(pt) を返す（無ければ None 相当）。"""
    m = re.search(r'<sheetFormatPr\b[^>]*>', sx)
    dc = dr = None
    if m:
        tag = m.group(0)
        mc = re.search(r'defaultColWidth="([\d.]+)"', tag)
        mr = re.search(r'defaultRowHeight="([\d.]+)"', tag)
        if mc:
            dc = float(mc.group(1))
        if mr:
            dr = float(mr.group(1))
    return dc, dr


def _col_widths(sx: str) -> dict[int, float]:
    """列インデックス(1始まり)→ 文字幅。<cols><col min max width> から。"""
    widths: dict[int, float] = {}
    cols_block = re.search(r'<cols>(.*?)</cols>', sx, re.DOTALL)
    if not cols_block:
        return widths
    for m in re.finditer(r'<col\b[^>]*?min="(\d+)"[^>]*?max="(\d+)"[^>]*?width="([\d.]+)"[^>]*?>', cols_block.group(1)):
        lo, hi, w = int(m.group(1)), int(m.group(2)), float(m.group(3))
        for c in range(lo, hi + 1):
            widths[c] = w
    return widths


def _row_heights(sx: str) -> dict[int, float]:
    """行番号(1始まり)→ 高さ(pt)。<row r ht> から。"""
    heights: dict[int, float] = {}
    for m in re.finditer(r'<row\b[^>]*?\br="(\d+)"[^>]*?\bht="([\d.]+)"', sx):
        heights[int(m.group(1))] = float(m.group(2))
    return heights


def _frame_px(sx: str, min_c: int, min_r: int, max_c: int, max_r: int) -> tuple[int, int]:
    """結合範囲の概算ピクセルサイズ（xlsx_writer._range_px と同じ換算）。"""
    cw = _col_widths(sx)
    rh = _row_heights(sx)
    dc, dr = _sheet_defaults(sx)
    w = 0.0
    for c in range(min_c, max_c + 1):
        width = cw.get(c, dc)
        w += width * 7.0 if width else _COL_PX
    h = 0.0
    for r in range(min_r, max_r + 1):
        height = rh.get(r, dr)
        h += height * 1.333 if height else _ROW_PX
    return int(w), int(h)


def set_cell_image(
    path: str,
    sheet: str,
    anchor: str,
    image_path: str,
    *,
    cell_range: Optional[str] = None,
    insert_rows: bool = False,
    width: Optional[int] = None,
    height: Optional[int] = None,
    max_dim: Optional[int] = None,
    col_off_px: int = 0,
    row_off_px: int = 0,
    output: Optional[str] = None,
) -> str:
    """既存内容（埋め込み画像含む）を保持して over-cell 画像を注入する。

    xlsx_writer.set_cell_image と同じ振る舞い:
    - cell_range 指定時はその範囲を結合し、画像を枠サイズへ**アスペクト維持**でフィット。
    - insert_rows=True で枠ぶんの行を先に挿入（既存は下へ）。
    - cell_range 省略時は anchor に width/height（または自然サイズ）で浮かせる。
    """
    from PIL import Image as PILImage

    src = downscale(image_path, max_dim)
    with PILImage.open(src) as im:
        iw, ih = im.size
    with open(src, "rb") as f:
        img_bytes = f.read()

    entries = _read_zip(path)
    sheet_xml = _sheet_xml_path(entries, sheet)
    draw_path, draw_rels = _drawing_for_sheet(entries, sheet_xml)

    # 配置先セルとサイズの決定
    if cell_range:
        min_c, min_r, max_c, max_r = range_boundaries(cell_range)
        nrows = max_r - min_r + 1
        if insert_rows:
            _shift_rows_down(entries, sheet_xml, draw_path, min_r, nrows)
        _remove_overlapping_merges(entries, sheet_xml, min_c, min_r, max_c, max_r)
        merge_ref = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        _add_merge(entries, sheet_xml, merge_ref)
        # 枠にアスペクト維持でフィット
        sx = entries[sheet_xml].decode("utf-8")
        fw, fh = _frame_px(sx, min_c, min_r, max_c, max_r)
        scale = min(fw / iw, fh / ih) if iw and ih else 1.0
        disp_w, disp_h = max(1, int(iw * scale)), max(1, int(ih * scale))
        col0, row0 = min_c - 1, min_r - 1
    else:
        row, col = coordinate_to_tuple(anchor)  # 返り値は (row, col)
        col0, row0 = col - 1, row - 1
        disp_w = width if width else iw
        disp_h = height if height else ih

    # media 追加
    midx = _next_media_index(entries)
    media_name = f"xl/media/image{midx}.png"
    entries[media_name] = img_bytes
    _ensure_png_content_type(entries)

    # drawing rels に画像 rel 追加
    rid = _add_rel(entries, draw_rels,
                   "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image",
                   f"../media/image{midx}.png")

    cx, cy = disp_w * _EMU, disp_h * _EMU
    pic_id = 100000 + midx
    dx = entries[draw_path].decode("utf-8")
    # 既存 drawing が xdr: 接頭辞か、デフォルト名前空間（接頭辞なし）かを検出して合わせる。
    # （Excel 由来=xdr:、openpyxl 由来=接頭辞なし。混在は不正 XML になるため必ず一致させる）
    p = "xdr:" if "<xdr:wsDr" in dx else ""
    a_ns = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
    r_ns = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
    anchor_xml = (
        f'<{p}oneCellAnchor>'
        f'<{p}from><{p}col>{col0}</{p}col><{p}colOff>{col_off_px * _EMU}</{p}colOff>'
        f'<{p}row>{row0}</{p}row><{p}rowOff>{row_off_px * _EMU}</{p}rowOff></{p}from>'
        f'<{p}ext cx="{cx}" cy="{cy}"/>'
        f'<{p}pic>'
        f'<{p}nvPicPr><{p}cNvPr id="{pic_id}" name="capture{midx}"/>'
        f'<{p}cNvPicPr><a:picLocks {a_ns} noChangeAspect="1"/></{p}cNvPicPr></{p}nvPicPr>'
        f'<{p}blipFill><a:blip {a_ns} {r_ns} r:embed="{rid}"/>'
        f'<a:stretch {a_ns}><a:fillRect/></a:stretch></{p}blipFill>'
        f'<{p}spPr><a:prstGeom {a_ns} prst="rect"><a:avLst/></a:prstGeom></{p}spPr>'
        f'</{p}pic><{p}clientData/></{p}oneCellAnchor>'
    )
    dx = _expand_self_closed_wsdr(dx)
    close = "</xdr:wsDr>" if p else "</wsDr>"
    dx = dx.replace(close, anchor_xml + close)
    entries[draw_path] = dx.encode("utf-8")

    dest = output or path
    _write_zip(entries, dest)
    return dest


def add_image_over_cells(
    path: str,
    sheet: str,
    anchor: str,
    image_path: str,
    *,
    cell_range: Optional[str] = None,
    max_dim: Optional[int] = None,
    output: Optional[str] = None,
) -> str:
    """後方互換ラッパ。oneCellAnchor で自然サイズ配置（cell_range 指定時は結合のみ追加）。"""
    return set_cell_image(
        path, sheet, anchor, image_path,
        cell_range=cell_range, insert_rows=False, max_dim=max_dim, output=output,
    )


# ---------------------------------------------------------------------------
# 画像上への注釈図形（赤枠+番号バッジ）の注入
#
# openpyxl は図形追加 API を持たず、load→save で図形(sp)を必ず落とすため、
# ここでも ZIP 直接編集で drawing XML に xdr:sp を追記する（画像注入と同方式）。
# ---------------------------------------------------------------------------

_EMU_PER_PT = 12700


def _pics_in_drawing(dx: str) -> list[dict]:
    """drawing XML 内の pic を持つアンカーを列挙する。

    返り値: {name, col, col_off, row, row_off, cx, cy, rid} のリスト。
    from（セルアンカー）と ext（表示 EMU 寸法）の両方を持つものだけを返す
    （absoluteAnchor や ext 無し twoCellAnchor は注釈の位置基準にできないため除外）。
    """
    pics: list[dict] = []
    for am in re.finditer(
        r"<(?:xdr:)?(oneCellAnchor|twoCellAnchor|absoluteAnchor)>.*?</(?:xdr:)?\1>",
        dx, re.DOTALL,
    ):
        block = am.group(0)
        if not re.search(r"<(?:xdr:)?pic>", block):
            continue
        name = re.search(r'<(?:xdr:)?cNvPr [^>]*name="([^"]*)"', block)
        frm = re.search(
            r"<(?:xdr:)?from>\s*<(?:xdr:)?col>(\d+)</(?:xdr:)?col>\s*"
            r"<(?:xdr:)?colOff>(-?\d+)</(?:xdr:)?colOff>\s*"
            r"<(?:xdr:)?row>(\d+)</(?:xdr:)?row>\s*"
            r"<(?:xdr:)?rowOff>(-?\d+)</(?:xdr:)?rowOff>",
            block,
        )
        ext = re.search(r'<(?:xdr:)?ext cx="(\d+)" cy="(\d+)"', block)
        rid = re.search(r'r:embed="(rId\d+)"', block)
        # Excel が再保存時に付ける絶対座標キャッシュ（<a:xfrm><a:off>）。
        # あれば注釈図形の xfrm を誤差ゼロで計算できる。
        off = re.search(r'<a:off x="(-?\d+)" y="(-?\d+)"/>', block)
        if not (frm and ext and rid):
            continue
        pics.append({
            "name": name.group(1) if name else "",
            "col": int(frm.group(1)), "col_off": int(frm.group(2)),
            "row": int(frm.group(3)), "row_off": int(frm.group(4)),
            "cx": int(ext.group(1)), "cy": int(ext.group(2)),
            "rid": rid.group(1),
            "off_x": int(off.group(1)) if off else None,
            "off_y": int(off.group(2)) if off else None,
        })
    return pics


def _anchor_origin_emu(sx: str, col0: int, row0: int) -> tuple[int, int]:
    """セル（0始まり col0,row0）左上の、シート原点からの絶対 EMU 座標。

    Excel 実式（列: round(文字幅×7)+5 px、行: pt×96/72 px）で換算した概算。
    pic に <a:off>（Excel の絶対座標キャッシュ）が無い場合のフォールバック用。
    """
    cw = _col_widths(sx)
    rh = _row_heights(sx)
    dc, dr = _sheet_defaults(sx)
    x = 0.0
    for c in range(1, col0 + 1):
        width = cw.get(c, dc)
        x += (round(width * 7.0) + 5) if width else _COL_PX
    y = 0.0
    for r in range(1, row0 + 1):
        height = rh.get(r, dr)
        y += height * 96.0 / 72.0 if height else _ROW_PX
    return round(x * _EMU), round(y * _EMU)


def _media_for_rid(entries: dict[str, bytes], draw_rels: str, rid: str) -> str:
    """drawing rels の rId から media エントリのパス（xl/media/imageN.png）を解決する。"""
    rels = entries.get(draw_rels, b"").decode("utf-8")
    rel = re.search(r'<Relationship\b[^>]*\bId="' + re.escape(rid) + r'"[^>]*/>', rels)
    if not rel:
        raise ValueError(f"画像 rel が見つかりません: {rid}")
    mt = re.search(r'Target="([^"]+)"', rel.group(0))
    if not mt:
        raise ValueError(f"画像 rel に Target がありません: {rid}")
    target = mt.group(1)
    if "../" in target:
        return "xl/" + target.split("../", 1)[1]
    return target.lstrip("/")


def add_annotations(
    path: str,
    sheet: str,
    annots: list[tuple[int, int, int, int, str]],
    *,
    name: Optional[str] = None,
    color: str = "FF0000",
    line_pt: float = 2.25,
    badge: bool = True,
    output: Optional[str] = None,
) -> str:
    """埋め込み画像の上へ赤枠矩形+番号バッジをネイティブ図形（xdr:sp）として注入する。

    annots: [(x, y, w, h, label), ...]。座標は**元画像（xl/media の実 PNG）のピクセル**。
    表示位置は「表示 EMU × 画像内比率」で厳密換算するため、列幅換算の近似誤差が乗らない。

    アンカーは absoluteAnchor（絶対 EMU）を使う。oneCellAnchor のオフセット合成だと
    Excel はセル幅/行高を超えるオフセットをセル境界へクランプして描画がズレる（実測）。
    絶対座標は pic の a:xfrm a:off（Excel の座標キャッシュ）があれば誤差ゼロ、無ければ
    列幅/行高からの概算。absoluteAnchor は行挿入に追従しないため、行操作の後に実行する。

    注意: openpyxl は load→save で図形を落とす（画像は保持）。set-text 等 openpyxl 系の
    書込より**後に**実行すること。消えた場合は再実行すればよい。
    """
    from PIL import Image as PILImage

    color = color.strip().lstrip("#").upper()
    if not re.fullmatch(r"[0-9A-F]{6}", color):
        raise ValueError(f"色は RRGGBB 16進6桁で指定してください: {color!r}")

    entries = _read_zip(path)
    sheet_xml = _sheet_xml_path(entries, sheet)
    found = _find_drawing(entries, sheet_xml)
    if not found:
        raise ValueError(f"シートに画像がありません: {sheet}")
    draw_path, draw_rels = found
    dx = entries[draw_path].decode("utf-8")

    pics = _pics_in_drawing(dx)
    if not pics:
        raise ValueError(f"シートに注釈可能な画像（セルアンカー+寸法つき）がありません: {sheet}")
    if name is not None:
        matches = [p for p in pics if p["name"] == name]
        if not matches:
            raise ValueError(
                f"画像が見つかりません: {name}"
                f"（候補: {', '.join(p['name'] or '(無名)' for p in pics)}）")
        pic = matches[0]
    elif len(pics) == 1:
        pic = pics[0]
    else:
        raise ValueError(
            "画像が複数あります。name で対象を指定してください"
            f"（候補: {', '.join(p['name'] or '(無名)' for p in pics)}）")

    media = _media_for_rid(entries, draw_rels, pic["rid"])
    with PILImage.open(io.BytesIO(entries[media])) as im:
        iw, ih = im.size

    # 画像左上の絶対 EMU 座標。absoluteAnchor の pos と spPr/a:xfrm の両方に使う
    # （Excel はアンカー、Quick Look 等の簡易レンダラは xfrm を見る。同値にして一致させる）。
    if pic["off_x"] is not None and pic["off_y"] is not None:
        pic_abs_x, pic_abs_y = pic["off_x"], pic["off_y"]
    else:
        ox, oy = _anchor_origin_emu(entries[sheet_xml].decode("utf-8"), pic["col"], pic["row"])
        pic_abs_x, pic_abs_y = ox + pic["col_off"], oy + pic["row_off"]

    from xml.sax.saxutils import escape

    p = "xdr:" if "<xdr:wsDr" in dx else ""
    a_ns = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
    next_id = max((int(i) for i in re.findall(r'<(?:xdr:)?cNvPr [^>]*id="(\d+)"', dx)),
                  default=1) + 1
    ln_emu = round(line_pt * _EMU_PER_PT)
    bd = max(round(0.07 * min(pic["cx"], pic["cy"])), 22 * _EMU)  # バッジ径(EMU)
    sz = max(600, round(bd / _EMU_PER_PT * 50))  # 番号フォント（1/100pt、径の約半分）
    # 既存 annot の連番に続ける（追記実行でも名前が衝突しない）
    seq = max((int(m) for m in re.findall(r'name="annot(\d+)"', dx)), default=0) + 1

    parts: list[str] = []
    for x, y, w, h, label in annots:
        dx_emu = x * pic["cx"] // iw
        dy_emu = y * pic["cy"] // ih
        w_emu = max(1, w * pic["cx"] // iw)
        h_emu = max(1, h * pic["cy"] // ih)
        abs_x = pic_abs_x + dx_emu
        abs_y = pic_abs_y + dy_emu
        parts.append(
            f'<{p}absoluteAnchor>'
            f'<{p}pos x="{abs_x}" y="{abs_y}"/>'
            f'<{p}ext cx="{w_emu}" cy="{h_emu}"/>'
            f'<{p}sp><{p}nvSpPr><{p}cNvPr id="{next_id}" name="annot{seq}"/><{p}cNvSpPr/></{p}nvSpPr>'
            f'<{p}spPr>'
            f'<a:xfrm {a_ns}><a:off x="{abs_x}" y="{abs_y}"/><a:ext cx="{w_emu}" cy="{h_emu}"/></a:xfrm>'
            f'<a:prstGeom {a_ns} prst="rect"><a:avLst/></a:prstGeom>'
            f'<a:noFill {a_ns}/>'
            f'<a:ln {a_ns} w="{ln_emu}"><a:solidFill><a:srgbClr val="{color}"/></a:solidFill></a:ln>'
            f'</{p}spPr></{p}sp><{p}clientData/></{p}absoluteAnchor>'
        )
        next_id += 1
        if badge:
            # 矩形の左上角にバッジ中心を重ねる（シート外に出ないよう 0 で下限クランプ）
            b_abs_x = max(0, abs_x - bd // 2)
            b_abs_y = max(0, abs_y - bd // 2)
            parts.append(
                f'<{p}absoluteAnchor>'
                f'<{p}pos x="{b_abs_x}" y="{b_abs_y}"/>'
                f'<{p}ext cx="{bd}" cy="{bd}"/>'
                f'<{p}sp><{p}nvSpPr><{p}cNvPr id="{next_id}" name="annot{seq}badge"/><{p}cNvSpPr/></{p}nvSpPr>'
                f'<{p}spPr>'
                f'<a:xfrm {a_ns}><a:off x="{b_abs_x}" y="{b_abs_y}"/><a:ext cx="{bd}" cy="{bd}"/></a:xfrm>'
                f'<a:prstGeom {a_ns} prst="ellipse"><a:avLst/></a:prstGeom>'
                f'<a:solidFill {a_ns}><a:srgbClr val="{color}"/></a:solidFill>'
                f'<a:ln {a_ns}><a:noFill/></a:ln></{p}spPr>'
                f'<{p}txBody><a:bodyPr {a_ns} anchor="ctr" lIns="0" tIns="0" rIns="0" bIns="0"/>'
                f'<a:lstStyle {a_ns}/>'
                f'<a:p {a_ns}><a:pPr algn="ctr"/>'
                f'<a:r><a:rPr lang="ja-JP" sz="{sz}" b="1">'
                f'<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill></a:rPr>'
                f'<a:t>{escape(label)}</a:t></a:r></a:p></{p}txBody>'
                f'</{p}sp><{p}clientData/></{p}absoluteAnchor>'
            )
            next_id += 1
        seq += 1

    dx = _expand_self_closed_wsdr(dx)
    close = f"</{p}wsDr>"
    entries[draw_path] = dx.replace(close, "".join(parts) + close).encode("utf-8")
    dest = output or path
    _write_zip(entries, dest)
    return dest


def remove_annotations(path: str, sheet: str, *, output: Optional[str] = None) -> tuple[str, int]:
    """add_annotations が注入した図形（cNvPr name=annot…）を除去する。返り値 (保存先, 除去数)。

    pic を含むアンカーは名前が annot… でも除去しない（画像の誤削除防止）。
    """
    entries = _read_zip(path)
    sheet_xml = _sheet_xml_path(entries, sheet)
    found = _find_drawing(entries, sheet_xml)
    dest = output or path
    if not found:
        _write_zip(entries, dest)
        return dest, 0
    draw_path, _ = found
    dx = entries[draw_path].decode("utf-8")
    removed = 0

    def drop(m: re.Match) -> str:
        nonlocal removed
        block = m.group(0)
        if (re.search(r'<(?:xdr:)?cNvPr [^>]*name="annot[^"]*"', block)
                and not re.search(r"<(?:xdr:)?pic>", block)):
            removed += 1
            return ""
        return block

    dx = re.sub(
        r"<(?:xdr:)?(oneCellAnchor|twoCellAnchor|absoluteAnchor)>.*?</(?:xdr:)?\1>",
        drop, dx, flags=re.DOTALL,
    )
    entries[draw_path] = dx.encode("utf-8")
    _write_zip(entries, dest)
    return dest, removed


# ── 図形（drawing）保全 ────────────────────────────────────────────────
# openpyxl は load→save の round-trip で drawing（画像・図形・コメントVML）を
# 落とす。テンプレの注意書きバナー等が消えるため、openpyxl 系コマンドの前後で
# 元 zip の drawing パーツを退避→再注入する。


def snapshot(path: str) -> dict[str, bytes]:
    """openpyxl 処理の前に元 zip の全パーツを退避する。"""
    return _read_zip(path)


def _sheet_name_map(entries: dict[str, bytes]) -> dict[str, str]:
    """シート名 -> xl/worksheets/sheetN.xml"""
    wb = entries["xl/workbook.xml"].decode("utf-8")
    rels = entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
    out: dict[str, str] = {}
    for m in re.finditer(r'<sheet\b[^>]*/>', wb):
        tag = m.group(0)
        mn = re.search(r'name="([^"]+)"', tag)
        mr = re.search(r'r:id="(rId\d+)"', tag)
        if not (mn and mr):
            continue
        rel = re.search(r'<Relationship\b[^>]*\bId="' + re.escape(mr.group(1)) + r'"[^>]*/>', rels)
        if not rel:
            continue
        mt = re.search(r'Target="([^"]+)"', rel.group(0))
        if not mt:
            continue
        t = mt.group(1).lstrip("/")
        out[mn.group(1)] = t if t.startswith("xl/") else "xl/" + t
    return out


def _next_index(entries: dict[str, bytes], prefix: str, suffix: str) -> int:
    idx = [int(m.group(1)) for n in entries
           if (m := re.fullmatch(re.escape(prefix) + r"(\d+)" + re.escape(suffix), n))]
    return (max(idx) + 1) if idx else 1


def _add_content_type(entries: dict[str, bytes], part: str, ctype: str) -> None:
    ct = entries["[Content_Types].xml"].decode("utf-8")
    if f'PartName="/{part}"' in ct:
        return
    ct = ct.replace("</Types>", f'<Override PartName="/{part}" ContentType="{ctype}"/></Types>')
    entries["[Content_Types].xml"] = ct.encode("utf-8")


_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _ensure_r_namespace(sx: str) -> str:
    """openpyxl が書いた worksheet には xmlns:r が無いことがある。r:id を使う前に宣言する。"""
    m = re.search(r"<worksheet\b[^>]*>", sx)
    if not m or 'xmlns:r=' in m.group(0):
        return sx
    tag = m.group(0)
    new_tag = tag[:-1].rstrip() + f' xmlns:r="{_R_NS}"' + ">"
    return sx[: m.start()] + new_tag + sx[m.end():]


def _insert_before_close(sx: str, element: str) -> str:
    """<worksheet> の正しい位置に要素を挿す（drawing は legacyDrawing の前）。"""
    sx = _ensure_r_namespace(sx)
    for anchor in ("<legacyDrawing", "<picture", "<oleObjects", "</worksheet>"):
        i = sx.find(anchor)
        if i != -1:
            return sx[:i] + element + sx[i:]
    return sx + element


def restore_drawings(orig: dict[str, bytes], dest: str,
                     except_sheets: Optional[set[str]] = None) -> int:
    """openpyxl 保存後のファイルに、元 zip の drawing/VML を復元する。

    復元先シートに既に drawing がある場合（gwex が新たに画像を貼ったシート）は
    そのまま残し、元の drawing は追加しない（新旧の取り違え防止）。
    戻り値は復元したシート数。
    """
    new = _read_zip(dest)
    orig_sheets = _sheet_name_map(orig)
    new_sheets = _sheet_name_map(new)
    restored = 0

    skip = except_sheets or set()
    for name, o_sheet in orig_sheets.items():
        if name in skip:
            continue
        n_sheet = new_sheets.get(name)
        if not n_sheet or o_sheet not in orig:
            continue

        o_base = o_sheet.split("/")[-1]
        o_rels = orig.get(f"xl/worksheets/_rels/{o_base}.rels", b"").decode("utf-8")
        if not o_rels:
            continue

        n_base = n_sheet.split("/")[-1]
        n_rels_path = f"xl/worksheets/_rels/{n_base}.rels"
        n_sx = new[n_sheet].decode("utf-8")
        touched = False

        # 1) drawing（画像・図形）
        md = re.search(r'Target="([^"]*drawings/(drawing\d+\.xml))"', o_rels)
        if md and f"xl/drawings/{md.group(2)}" in orig:
            o_draw = f"xl/drawings/{md.group(2)}"
            o_name = md.group(2)
            dx = orig[o_draw].decode("utf-8")

            # openpyxl は「空の drawing の殻」だけ残すことがある（rel は張られているが中身が空）。
            # 中身のある drawing が既にあるなら触らない（gwex が新たに貼った画像を守る）。
            existing = _find_drawing(new, n_sheet)
            target_draw = None
            merge_shapes = None
            if existing:
                nx = new.get(existing[0], b"").decode("utf-8")
                if "Anchor" not in nx:
                    # 空殻（openpyxl が rel だけ残した）→ 元の中身をそのまま復元
                    target_draw = existing[0]
                # 中身のある drawing（openpyxl が画像だけ引き継いだ等）は触らない。
                # 図形（annotate-image の赤枠等）はここでは復元しない —— 二重注入で
                # 画像/図形が重複するため。注釈は従来どおりワークフロー最終段で実行する。
            else:
                k = _next_index(new, "xl/drawings/drawing", ".xml")
                target_draw = f"xl/drawings/drawing{k}.xml"

            if merge_shapes:
                path_, nx, lost = merge_shapes
                nx = _ensure_drawing_namespaces(_expand_self_closed_wsdr(nx))
                block = _renumber_ids("".join(lost), _max_shape_id(nx))
                close = "</xdr:wsDr>" if "</xdr:wsDr>" in nx else "</wsDr>"
                new[path_] = nx.replace(close, block + close).encode("utf-8")
                touched = True

            if target_draw:
                # drawing の rels（media 参照）をコピーしつつ media を採番し直す
                o_draw_rels = f"xl/drawings/_rels/{o_name}.rels"
                t_name = target_draw.split("/")[-1]
                if o_draw_rels in orig:
                    drels = orig[o_draw_rels].decode("utf-8")
                    for mm in re.finditer(r'Target="([^"]*media/(image[\w.]+))"', drels):
                        o_media = f"xl/media/{mm.group(2)}"
                        if o_media not in orig:
                            continue
                        ext = mm.group(2).rsplit(".", 1)[-1]
                        j = _next_index(new, "xl/media/image", f".{ext}")
                        new_media = f"xl/media/image{j}.{ext}"
                        new[new_media] = orig[o_media]
                        drels = drels.replace(mm.group(1), f"../media/image{j}.{ext}")
                        _add_content_type(new, new_media,
                                          f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}")
                    new[f"xl/drawings/_rels/{t_name}.rels"] = drels.encode("utf-8")

                new[target_draw] = dx.encode("utf-8")
                _add_content_type(new, target_draw,
                                  "application/vnd.openxmlformats-officedocument.drawing+xml")
                if not existing:
                    rid = _add_rel(new, n_rels_path,
                                   "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
                                   f"../drawings/{t_name}")
                    n_sx = _insert_before_close(n_sx, f'<drawing r:id="{rid}"/>')
                touched = True

        # 2) legacyDrawing（コメントの VML）
        mv = re.search(r'Target="([^"]*drawings/(vmlDrawing\d+\.vml))"', o_rels)
        if mv and "<legacyDrawing" not in n_sx:
            o_vml = f"xl/drawings/{mv.group(2)}"
            if o_vml in orig:
                k = _next_index(new, "xl/drawings/vmlDrawing", ".vml")
                new_vml = f"xl/drawings/vmlDrawing{k}.vml"
                new[new_vml] = orig[o_vml]
                ct = new["[Content_Types].xml"].decode("utf-8")
                if 'Extension="vml"' not in ct:
                    ct = ct.replace("<Types", "<Types", 1).replace(
                        "</Types>",
                        '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/></Types>')
                    new["[Content_Types].xml"] = ct.encode("utf-8")
                rid = _add_rel(new, n_rels_path,
                               "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing",
                               f"../drawings/vmlDrawing{k}.vml")
                n_sx = _insert_before_close(n_sx, f'<legacyDrawing r:id="{rid}"/>')
                touched = True

        if touched:
            new[n_sheet] = n_sx.encode("utf-8")
            restored += 1

    if restored:
        _write_zip(new, dest)
    return restored
