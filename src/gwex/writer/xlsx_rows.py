"""ローカル .xlsx の行削除・挿入（openpyxl）。privileged。

openpyxl の `Worksheet.delete_rows` / `insert_rows` はセル値しかシフトせず、
結合セル・行の高さ・条件付き書式・data validation を据え置いたままにする。
本モジュールはそれらも正しくシフトしながら行削除・挿入する。

主用途:
- テンプレ複製時の「指示行（行1）」削除（結合/行高/条件付き書式/dv の -1 シフト込み）
- 未使用サンプル行・ブロックの削除
- 修正内容の依頼行追加（answerと同じ行挿入方式・書式はテンプレート行からコピー）
"""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries


def _shift_range(min_c, min_r, max_c, max_r, start, count):
    """行 [start, start+count-1] を削除したときの結合/書式範囲の新しい行境界を返す。
    完全に削除範囲へ収まる場合は None（範囲消滅）。"""
    end = start + count - 1
    if max_r < start:                      # 完全に上 → 据置
        new_min_r, new_max_r = min_r, max_r
    elif min_r > end:                      # 完全に下 → -count
        new_min_r, new_max_r = min_r - count, max_r - count
    else:                                  # 削除範囲と交差
        upper = start - min_r if min_r < start else 0      # 削除範囲より上に残る行数
        lower = max_r - end if max_r > end else 0          # 削除範囲より下に残る行数
        if upper == 0 and lower == 0:
            return None                    # 全消滅
        new_min_r = min_r if min_r < start else start
        new_max_r = (max_r - count) if max_r > end else (start - 1)
    return new_min_r, new_max_r


def _shift_sqref(sqref, start, count):
    """空白区切りの sqref（複数範囲可）を行削除に合わせてシフト。残った範囲文字列を返す。"""
    parts = []
    for p in str(sqref).split():
        min_c, min_r, max_c, max_r = range_boundaries(p)
        shifted = _shift_range(min_c, min_r, max_c, max_r, start, count)
        if shifted is None:
            continue
        nmin_r, nmax_r = shifted
        parts.append(
            f"{get_column_letter(min_c)}{nmin_r}:{get_column_letter(max_c)}{nmax_r}"
        )
    return " ".join(parts)


def _delete_rows_full(ws, start: int, count: int) -> None:
    """結合/行高/条件付き書式/dv を保ったまま行を削除する。"""
    # 1) 結合セルを退避して全 unmerge
    merged = [(range_boundaries(str(m)), get_column_letter(range_boundaries(str(m))[0]),
               get_column_letter(range_boundaries(str(m))[2])) for m in list(ws.merged_cells.ranges)]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    # 2) 条件付き書式を退避してクリア
    cf_rules = [(rng.sqref, [copy(r) for r in rules])
                for rng, rules in list(ws.conditional_formatting._cf_rules.items())]
    ws.conditional_formatting._cf_rules.clear()

    # 3) data validation を退避してクリア
    dvs = [(str(dv.sqref), dv) for dv in list(ws.data_validations.dataValidation)]
    ws.data_validations.dataValidation.clear()

    # 4) 行の高さを退避
    heights = {r: ws.row_dimensions[r].height
               for r in ws.row_dimensions if ws.row_dimensions[r].height is not None}

    # 5) 値のシフト（openpyxl 標準）
    ws.delete_rows(start, count)

    # 6) 行高を再設定
    end = start + count - 1
    for r in sorted(heights):
        if r < start:
            ws.row_dimensions[r].height = heights[r]
        elif r > end:
            ws.row_dimensions[r - count].height = heights[r]

    # 7) 結合を再 merge
    for (min_c, min_r, max_c, max_r), col_l, col_r in merged:
        shifted = _shift_range(min_c, min_r, max_c, max_r, start, count)
        if shifted is None:
            continue
        nmin_r, nmax_r = shifted
        ws.merge_cells(f"{col_l}{nmin_r}:{col_r}{nmax_r}")

    # 8) 条件付き書式を再登録
    for sqref, rules in cf_rules:
        new_sqref = _shift_sqref(sqref, start, count)
        if not new_sqref:
            continue
        for rule in rules:
            ws.conditional_formatting.add(new_sqref, rule)

    # 9) data validation を再登録
    for sqref, dv in dvs:
        new_sqref = _shift_sqref(sqref, start, count)
        if not new_sqref:
            continue
        dv.sqref = new_sqref
        ws.add_data_validation(dv)


def _insert_shift_range(min_c, min_r, max_c, max_r, at, count):
    """行 at の直前に count 行を挿入したときの範囲の新しい行境界を返す。"""
    if max_r < at:                       # 完全に挿入位置より上 → 据置
        return min_r, max_r
    new_min_r = min_r + count if min_r >= at else min_r
    new_max_r = max_r + count            # max_r >= at は常に +count
    return new_min_r, new_max_r


def _insert_shift_sqref(sqref, at, count):
    """空白区切りの sqref を行挿入に合わせてシフト。"""
    parts = []
    for p in str(sqref).split():
        min_c, min_r, max_c, max_r = range_boundaries(p)
        nmin_r, nmax_r = _insert_shift_range(min_c, min_r, max_c, max_r, at, count)
        parts.append(
            f"{get_column_letter(min_c)}{nmin_r}:{get_column_letter(max_c)}{nmax_r}"
        )
    return " ".join(parts)


def _insert_rows_full(ws, at: int, count: int, template_row: Optional[int] = None) -> None:
    """結合/行高/条件付き書式/dv を保ったまま行を挿入する。

    at: 挿入位置（この行の直前に count 行を挿入。1始まり）
    template_row: この行のセル書式・行高を新行にコピー（修正内容の依頼行追加用）
    """
    # 1) 結合セルを退避して全 unmerge
    merged = []
    for m in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = range_boundaries(str(m))
        merged.append((min_c, min_r, max_c, max_r,
                        get_column_letter(min_c), get_column_letter(max_c)))
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    # 2) 条件付き書式を退避してクリア
    cf_rules = [(rng.sqref, [copy(r) for r in rules])
                for rng, rules in list(ws.conditional_formatting._cf_rules.items())]
    ws.conditional_formatting._cf_rules.clear()

    # 3) data validation を退避してクリア
    dvs = [(str(dv.sqref), dv) for dv in list(ws.data_validations.dataValidation)]
    ws.data_validations.dataValidation.clear()

    # 4) 行の高さを退避
    heights = {r: ws.row_dimensions[r].height
               for r in ws.row_dimensions if ws.row_dimensions[r].height is not None}

    # 5) テンプレート行のセル書式・行高・結合範囲を退避
    template_styles = {}
    template_height = None
    template_merges = []   # テンプレート行を含む結合の列範囲 (col_l, col_r)
    if template_row is not None:
        for c in range(1, ws.max_column + 1):
            src = ws.cell(template_row, c)
            if src.has_style:
                template_styles[c] = copy(src._style)
        template_height = heights.get(template_row)
        # テンプレート行を含む結合の列範囲を収集（行挿入前・挿入前の行番号で）
        for min_c, min_r, max_c, max_r, col_l, col_r in merged:
            if min_r <= template_row <= max_r:
                template_merges.append((col_l, col_r))

    # 6) 値のシフト（openpyxl 標準）
    ws.insert_rows(at, count)

    # 7) 行高を再設定（at 以降の行を +count シフト）
    for r in sorted(heights, reverse=True):
        if r >= at:
            ws.row_dimensions[r + count].height = heights[r]
        # r < at は据置き（openpyxl が insert_rows で値のみシフト済み）

    # 8) テンプレート行の書式を新行にコピー
    if template_row is not None:
        for new_r in range(at, at + count):
            for c, style in template_styles.items():
                ws.cell(new_r, c)._style = copy(style)
            if template_height is not None:
                ws.row_dimensions[new_r].height = template_height

    # 9) 結合を再 merge（シフト後座標で）
    for min_c, min_r, max_c, max_r, col_l, col_r in merged:
        nmin_r, nmax_r = _insert_shift_range(min_c, min_r, max_c, max_r, at, count)
        ws.merge_cells(f"{col_l}{nmin_r}:{col_r}{nmax_r}")

    # 9b) テンプレート行の結合列範囲を新行にもコピー
    for col_l, col_r in template_merges:
        for new_r in range(at, at + count):
            try:
                ws.merge_cells(f"{col_l}{new_r}:{col_r}{new_r}")
            except Exception:
                pass  # 既に同範囲の結合がある場合はスキップ

    # 10) 条件付き書式を再登録
    for sqref, rules in cf_rules:
        new_sqref = _insert_shift_sqref(sqref, at, count)
        if not new_sqref:
            continue
        for rule in rules:
            ws.conditional_formatting.add(new_sqref, rule)

    # 11) data validation を再登録
    for sqref, dv in dvs:
        new_sqref = _insert_shift_sqref(sqref, at, count)
        if not new_sqref:
            continue
        dv.sqref = new_sqref
        ws.add_data_validation(dv)


def insert_rows(path: str, sheet: str, at: int, count: int,
                template_row: Optional[int] = None,
                *, output: Optional[str] = None) -> str:
    """指定シートの行 at の直前に count 行を挿入（結合/行高/条件付き書式/dv をシフト）。
    template_row を指定するとその行の書式（行高・セル書式）を新行にコピーする。
    """
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("insert_rows はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    ws = wb[sheet]
    _insert_rows_full(ws, at, count, template_row=template_row)
    dest = output or path
    wb.save(dest)
    return dest


def _copy_cell_format(src_cell, dst_cell) -> None:
    """セルの書式（font/fill/border/alignment/number_format）を個別属性コピーする。

    `_style = copy(src._style)` は StyleArray の index が workbook 固有のため
    クロス workbook では別スタイルを参照してしまう。個別属性コピーで確実に転写する。
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers

    # Font
    sf = src_cell.font
    dst_cell.font = Font(
        name=sf.name, size=sf.size, bold=sf.bold, italic=sf.italic,
        underline=sf.underline, strike=sf.strike, color=copy(sf.color),
    )
    # Fill
    spf = src_cell.fill
    dst_cell.fill = PatternFill(
        fill_type=spf.fill_type,
        fgColor=copy(spf.fgColor),
        bgColor=copy(spf.bgColor),
    )
    # Border
    def _side(s):
        return Side(border_style=s.border_style, color=copy(s.color)) if s else Side()
    sb = src_cell.border
    dst_cell.border = Border(
        left=_side(sb.left), right=_side(sb.right),
        top=_side(sb.top), bottom=_side(sb.bottom),
        diagonal=_side(sb.diagonal),
        diagonal_direction=sb.diagonal_direction,
        outline=sb.outline,
    )
    # Alignment
    sa = src_cell.alignment
    dst_cell.alignment = Alignment(
        horizontal=sa.horizontal, vertical=sa.vertical,
        wrap_text=sa.wrap_text, text_rotation=sa.text_rotation,
        shrink_to_fit=sa.shrink_to_fit, indent=sa.indent,
    )
    # Number format
    dst_cell.number_format = src_cell.number_format


def copy_row_format(
    src_path: str, src_sheet: str, src_rows: list,
    dst_path: str, dst_sheet: str, dst_rows: list,
    *, output: Optional[str] = None,
) -> str:
    """ソース xlsx の指定行の書式（行高・セル書式）をターゲット xlsx の指定行にコピーする。
    src_rows と dst_rows は同じ長さ。値はコピーしない（書式のみ）。
    クロス workbook コピーにも対応（個別属性コピー方式）。
    """
    if len(src_rows) != len(dst_rows):
        raise ValueError("src_rows と dst_rows の長さが一致しません。")
    wb_src = load_workbook(src_path)
    ws_src = wb_src[src_sheet]
    wb_dst = load_workbook(dst_path)
    ws_dst = wb_dst[dst_sheet]
    max_col = max(ws_src.max_column, ws_dst.max_column)
    for sr, dr in zip(src_rows, dst_rows):
        src_h = ws_src.row_dimensions[sr].height
        if src_h is not None:
            ws_dst.row_dimensions[dr].height = src_h
        for c in range(1, max_col + 1):
            src_cell = ws_src.cell(sr, c)
            dst_cell = ws_dst.cell(dr, c)
            if src_cell.has_style:
                _copy_cell_format(src_cell, dst_cell)
    dest = output or dst_path
    wb_dst.save(dest)
    return dest


def set_alignment(
    path: str, sheet: str, cells: list[str],
    *, horizontal: Optional[str] = None, vertical: Optional[str] = None,
    wrap_text: Optional[bool] = None, output: Optional[str] = None,
) -> str:
    """指定セルのアライメントを部分的に上書きする（未指定の属性は現状維持）。

    cells は ["C6", "C7", ...] または ["C6:C11"] 形式のリスト。
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import rows_from_range

    wb = load_workbook(path)
    ws = wb[sheet]

    target_cells: list[tuple[int, int]] = []
    for spec in cells:
        if ":" in spec:
            for row_cells in rows_from_range(spec):
                for addr in row_cells:
                    from openpyxl.utils.cell import coordinate_to_tuple
                    r, c = coordinate_to_tuple(addr)
                    target_cells.append((r, c))
        else:
            from openpyxl.utils.cell import coordinate_to_tuple
            r, c = coordinate_to_tuple(spec)
            target_cells.append((r, c))

    for r, c in target_cells:
        cell = ws.cell(r, c)
        old = cell.alignment
        cell.alignment = Alignment(
            horizontal=horizontal if horizontal is not None else old.horizontal,
            vertical=vertical if vertical is not None else old.vertical,
            wrap_text=wrap_text if wrap_text is not None else old.wrap_text,
            text_rotation=old.text_rotation,
            shrink_to_fit=old.shrink_to_fit,
            indent=old.indent,
        )

    dest = output or path
    wb.save(dest)
    return dest


def delete_rows(path: str, sheet: str, start: int, count: int,
                *, output: Optional[str] = None) -> str:
    """指定シートの行 [start, start+count-1] を削除（結合/行高/条件付き書式/dv をシフト）。"""
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("delete_rows はローカル .xlsx のみ対応です（Google スプレッドシート不可）。")
    wb = load_workbook(path)
    ws = wb[sheet]
    _delete_rows_full(ws, start, count)
    dest = output or path
    wb.save(dest)
    return dest


def clear_data_region(path: str, sheet: str, start: int, count: int,
                      *, left_col: str = "B", right_col: str = "S",
                      output: Optional[str] = None) -> str:
    """データ領域 [start, start+count-1] × [left_col, right_col] の結合を解除し値を消す。

    write-testspec の前処理用。テンプレのサンプル行に残る結合（E11:E12 等）と
    プレースホルダ（{C11}…）が記入を壊す（結合セルへの書込が左上にリダイレクトされる）ため、
    記入前にこの領域だけを更地化する。罫線・行高・書式は保持する（値と結合のみ操作）。
    """
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("clear_data_region はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    ws = wb[sheet]
    lc = column_index_from_string(left_col)
    rc = column_index_from_string(right_col)
    end = start + count - 1
    for m in list(ws.merged_cells.ranges):
        mc1, mr1, mc2, mr2 = range_boundaries(str(m))
        if mr2 >= start and mr1 <= end and mc2 >= lc and mc1 <= rc:
            ws.unmerge_cells(str(m))
    for r in range(start, end + 1):
        for c in range(lc, rc + 1):
            ws.cell(r, c).value = None
    dest = output or path
    wb.save(dest)
    return dest


def clear_images(path: str, sheet: str, *, output: Optional[str] = None) -> tuple[str, int]:
    """シートの埋め込み画像を全削除する（set-image の逆。スプシ変換前の画像なし版作成用）。
    返り値: (出力先, 削除した画像数)。"""
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("clear_images はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    ws = wb[sheet]
    n = len(ws._images)
    ws._images = []
    dest = output or path
    wb.save(dest)
    return dest, n


def extract_sheet(path: str, sheet: str, output: str) -> str:
    """対象シートだけを残した .xlsx を別ファイルに書き出す（Quick Look 目視用。
    qlmanage は先頭シートしか描画しないため、2枚目以降のシートを目視するときに使う）。"""
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("extract_sheet はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"シートがありません: {sheet}")
    for name in [s for s in wb.sheetnames if s != sheet]:
        del wb[name]
    wb.save(output)
    return output


def delete_instruction_row(path: str, sheet: Optional[str] = None,
                           *, marker: str = "{A$1}", output: Optional[str] = None) -> str:
    """行1が指示行（A1 が marker で始まる）のシートで行1を削除する。
    sheet 省略時は指示行を持つ全シートを自動検出して処理する。"""
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("delete_instruction_row はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    targets = [sheet] if sheet else list(wb.sheetnames)
    done = []
    for name in targets:
        ws = wb[name]
        a1 = ws.cell(1, 1).value
        if isinstance(a1, str) and a1.startswith(marker):
            _delete_rows_full(ws, 1, 1)
            done.append(name)
        elif sheet:  # 明示指定なのに指示行が無い → エラーにせず警告的にスキップ
            raise ValueError(f"シート '{name}' の A1 は指示行（{marker}…）ではありません。")
    dest = output or path
    wb.save(dest)
    return dest, done


def add_sheet(path: str, sheet: str, *, index: Optional[int] = None,
              exist_ok: bool = False, output: Optional[str] = None) -> str:
    """ローカル .xlsx に空シートを追加する（レビュー結果シート等の追記用）。

    index を渡すとその位置（0始まり）に挿入する。省略時は末尾。
    既に同名シートがある場合、exist_ok=False なら例外、True なら何もしない。
    """
    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("add_sheet はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    if sheet in wb.sheetnames:
        if not exist_ok:
            raise ValueError(f"同名シートが既にあります: {sheet}")
        dest = output or path
        wb.save(dest)
        return dest
    wb.create_sheet(title=sheet, index=index)
    dest = output or path
    wb.save(dest)
    return dest


def set_table(path: str, sheet: str, start_cell: str, rows: list[list], *,
              header: bool = True, output: Optional[str] = None) -> str:
    """2次元配列を start_cell を左上として一括書き込みする（罫線つき・先頭行は見出し）。

    set-text を何十回も叩く代わりに使う。xlsx のみ。
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if Path(path).suffix.lower() != ".xlsx":
        raise ValueError("set_table はローカル .xlsx のみ対応です。")
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"シートがありません: {sheet}")
    ws = wb[sheet]

    col0, row0 = _split_cell(start_cell)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="E8EDF3")

    for r, line in enumerate(rows):
        for c, val in enumerate(line):
            cell = ws.cell(row=row0 + r, column=col0 + c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if header and r == 0:
                cell.font = Font(bold=True)
                cell.fill = head_fill

    dest = output or path
    wb.save(dest)
    return dest


def _split_cell(cell: str) -> tuple[int, int]:
    letters = "".join(ch for ch in cell if ch.isalpha())
    digits = "".join(ch for ch in cell if ch.isdigit())
    return column_index_from_string(letters), int(digits)
