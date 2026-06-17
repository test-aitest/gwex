"""ローカル .xlsx への書き戻し（openpyxl）。privileged。

既定は in-place 上書き。`output` 指定時のみ別名保存。
"""

from __future__ import annotations

from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from gwex.domains.model import DEFAULT_MAPPING, MappingConfig

_DATA_ROLES = (
    "screen_name", "medium_category", "small_category",
    "test_no", "verification_content", "execution_steps",
)


def _anchor(ws, cell: str) -> str:
    """cell が結合範囲内なら左上アンカーに解決する（結合セルは左上のみ書込可）。"""
    row, col = coordinate_to_tuple(cell)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell


def set_cell_text(
    path: str, sheet: str, cell: str, text: str, *, output: Optional[str] = None
) -> str:
    wb = load_workbook(path)
    ws = wb[sheet]
    ws[_anchor(ws, cell)] = text
    dest = output or path
    wb.save(dest)
    return dest


def set_row_height(
    path: str, sheet: str, row: int, height: float, *, output: Optional[str] = None
) -> str:
    """指定行の行高を設定する。"""
    wb = load_workbook(path)
    ws = wb[sheet]
    ws.row_dimensions[row].height = height
    dest = output or path
    wb.save(dest)
    return dest


def set_merge_cells(
    path: str, sheet: str, ranges: list[str], *, output: Optional[str] = None
) -> str:
    """指定セル範囲を結合する（オーバーラップする既存結合を全て解除してから再結合）。"""
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(path)
    ws = wb[sheet]
    for rng in ranges:
        min_c, min_r, max_c, max_r = range_boundaries(rng)
        # オーバーラップする既存結合を全て解除
        overlapping = [
            r for r in list(ws.merged_cells.ranges)
            if not (r.max_col < min_c or r.min_col > max_c or
                    r.max_row < min_r or r.min_row > max_r)
        ]
        for r in overlapping:
            ws.unmerge_cells(str(r))
        ws.merge_cells(rng)
    dest = output or path
    wb.save(dest)
    return dest


def write_cells(
    path: str, sheet: str, assignments: list[tuple[str, object]], *, output: Optional[str] = None
) -> str:
    """[(セル, 値), ...] を一括書込（構造書き戻し用）。"""
    wb = load_workbook(path)
    ws = wb[sheet]
    for cell, value in assignments:
        ws[_anchor(ws, cell)] = value
    dest = output or path
    wb.save(dest)
    return dest


def last_data_row(path: str, sheet: str, mapping: Optional[MappingConfig] = None) -> int:
    """testspec のデータ列に値がある最終行（1始まり）を返す。データ無しなら data_start_row-1。"""
    cfg = mapping or DEFAULT_MAPPING
    cols = [i for i in (cfg.column_index(r) for r in _DATA_ROLES) if i is not None]
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    last = cfg.data_start_row - 1
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < cfg.data_start_row:
            continue
        if any(c < len(row) and row[c] is not None for c in cols):
            last = r_idx
    return last
