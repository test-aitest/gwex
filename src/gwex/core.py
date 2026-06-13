"""fetcher（取得）と parser（解析）を調停するオーケストレーション層。

ここは pure ではない（fetcher を import してよい）。pure 層は parsers/serializers/ir のみ。
解析は既定で sandbox（network/fs-write deny の別プロセス）経由で行う。
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from gwex import parser_entry, sandbox
from gwex.fetcher import local
from gwex.ir import Document
from gwex.serializers import json as json_ser
from gwex.serializers import markdown as md_ser


def to_document(source: str, *, use_sandbox: bool = True) -> Document:
    """ソース指定（Google URL / ローカル OOXML パス）を IR Document に変換する。

    fetcher だけがネットワーク/ディスク I/O を行い、解析は raw（bytes / JSON）に対して
    confined parser プロセスで実行する。
    """
    from gwex.fetcher import google

    file_id = google.detect(source)
    if file_id is not None:
        kind, raw = google.fetch(file_id)
        return _parse(kind, raw, use_sandbox)

    path = Path(source)
    if path.exists():
        kind = local.detect_source(path)
        if kind is None:
            raise ValueError(f"未対応のファイル形式です: {path.suffix}")
        return _parse(kind, local.read_bytes(path), use_sandbox)

    raise ValueError(f"ソースを解決できません: {source}")


def _parse(kind: str, raw: bytes, use_sandbox: bool) -> Document:
    if use_sandbox and sandbox.available():
        return sandbox.run_parser_confined(kind, raw)
    return parser_entry._parse(kind, raw)


def serialize(doc: Document, to: str) -> str:
    if to == "md":
        return md_ser.render(doc)
    if to == "json":
        return json_ser.render(doc)
    raise ValueError(f"未対応の出力形式: {to}（md|json）")


def convert(source: str, to: str = "md", *, use_sandbox: bool = True) -> str:
    return serialize(to_document(source, use_sandbox=use_sandbox), to)


# --------------------------------------------------------------------------
# テスト仕様（ドメイン抽出）
# --------------------------------------------------------------------------


def _xlsx_bytes(source: str) -> bytes:
    """source（ローカル .xlsx / Google URL）から xlsx バイト列を得る。"""
    from gwex.fetcher import google

    file_id = google.detect(source)
    if file_id is not None:
        # ネイティブ Google シートは Drive で xlsx にエクスポート、アップロード xlsx は実バイト取得
        return google.fetch_xlsx(file_id)
    path = Path(source)
    if not path.exists():
        raise ValueError(f"ソースを解決できません: {source}")
    if local.detect_source(path) != "xlsx":
        raise ValueError("testspec 抽出は .xlsx のみ対応です。")
    return local.read_bytes(path)


def to_testspec(source: str, sheet: str, mapping=None):
    from gwex.domains import testspec

    return testspec.extract(_xlsx_bytes(source), sheet, mapping)


def serialize_testspec(spec, to: str) -> str:
    from gwex.serializers import testspec_json, testspec_md

    if to == "md":
        return testspec_md.render(spec)
    if to == "json":
        return testspec_json.render(spec)
    raise ValueError(f"未対応の出力形式: {to}（md|json）")


def write_testspec(
    source_json: str,
    target: str,
    sheet: str,
    mapping=None,
    *,
    append: bool = False,
    dedup: bool = False,
    apply_format: bool = False,
    clear_rows: Optional[int] = None,
    output=None,
) -> str:
    """TestSpec(JSON) を target（ローカル .xlsx / Google URL）のシートに展開記入する。

    append: 既存データの下に追記。dedup: 既存と完全一致するケースを除外し冪等に。
    append/dedup 時は採番を既存グループの最大Noから継続する。
    apply_format: 追記ブロックに既存テンプレ体裁（罫線/縦結合/親番号）を後付けする（ローカル .xlsx のみ）。
    """
    from gwex.domains import testspec_writer
    from gwex.domains.model import DEFAULT_MAPPING, TestSpec
    from gwex.fetcher import google

    spec = TestSpec.model_validate_json(source_json)
    is_google = google.detect(target) is not None

    existing = None
    if dedup or append:
        try:
            existing = to_testspec(target, sheet, mapping)  # native gsheet 等は ValueError
        except ValueError:
            existing = None
    if existing is not None and (dedup or append):
        spec = _prepare_spec(spec, existing, mapping, drop_dups=dedup)

    start_row = None
    if append:
        start_row = _last_data_row(target, sheet, mapping, is_google) + 1

    if clear_rows and not is_google and not append:
        from gwex.writer import xlsx_rows

        cfg0 = mapping or DEFAULT_MAPPING
        target = xlsx_rows.clear_data_region(
            target, sheet, cfg0.data_start_row, clear_rows,
            left_col=cfg0.box_left_col or "A", right_col=cfg0.box_right_col or "Z",
            output=output,
        )
        output = None  # クリア結果を target に取り込んだので以降は in-place

    asgn = testspec_writer.assignments(spec, mapping, start_row=start_row)
    if is_google:
        from gwex.writer import gsheet_writer

        return gsheet_writer.write_cells(target, sheet, asgn)
    from gwex.writer import xlsx_writer

    dest = xlsx_writer.write_cells(target, sheet, asgn, output=output)

    if apply_format:
        from gwex.writer import xlsx_format

        cfg = mapping or DEFAULT_MAPPING
        screen_start_no = (_max_screen_no(target, sheet, cfg) + 1) if existing is not None else 1
        plan = testspec_writer.layout(
            spec, cfg,
            start_row=(start_row if start_row is not None else cfg.data_start_row),
            screen_start_no=screen_start_no,
        )
        xlsx_format.format_testspec_block(dest, sheet, plan, cfg)
    return dest


def create_section(
    target: str,
    sheet: str,
    top_row: int,
    title: str,
    before_image=None,
    after_image=None,
    *,
    left_cols=("C", "L"),
    split_col: str = "H",
    scale: float = 0.8,
    output=None,
) -> str:
    """before/after セクション（枠/青背景/中央/画像80%中央）を作る。target で Excel/Google 振り分け。"""
    from gwex.fetcher import google

    if google.detect(target) is not None:
        from gwex.writer import gsheet_format

        return gsheet_format.create_before_after_section(
            target, sheet, top_row, title, before_image, after_image,
            left_cols=left_cols, split_col=split_col, scale=scale,
        )
    from gwex.writer import xlsx_format

    return xlsx_format.create_before_after_section(
        target, sheet, top_row, title, before_image, after_image,
        left_cols=left_cols, split_col=split_col, output=output,
    )


def _max_screen_no(target: str, sheet: str, cfg) -> int:
    """既存の画面No列（number_columns['screen_name']）の最大整数を返す。無ければ 0。"""
    col = (cfg.number_columns or {}).get("screen_name")
    if not col:
        return 0
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(target, read_only=True)
    ws = wb[sheet]
    ci = column_index_from_string(col)
    mx = 0
    for r_idx, row in enumerate(ws.iter_rows(min_col=ci, max_col=ci, values_only=True), start=1):
        if r_idx < cfg.data_start_row:
            continue
        v = row[0]
        if isinstance(v, int):
            mx = max(mx, v)
        elif isinstance(v, str) and v.strip().isdigit():
            mx = max(mx, int(v.strip()))
    return mx


def _last_data_row(target: str, sheet: str, mapping, is_google: bool) -> int:
    if is_google:
        from gwex.writer import gsheet_writer

        return gsheet_writer.last_data_row(target, sheet, mapping)
    from gwex.writer import xlsx_writer

    return xlsx_writer.last_data_row(target, sheet, mapping)


def _prepare_spec(spec, existing, mapping, *, drop_dups: bool):
    """既存と突き合わせ、(drop_dups なら)完全一致を除外し、採番を既存グループ最大Noから継続する。"""
    from gwex.domains.model import Group, Screen, TestSpec
    from gwex.domains.testspec_merge import case_key, group_key

    ex_keys = set()
    ex_max: dict[tuple, int] = {}
    for sc in existing.screens:
        for g in sc.groups:
            gk = group_key(sc.screen_name, g.medium_category, g.small_category)
            for c in g.cases:
                ex_keys.add(case_key(sc.screen_name, g.medium_category, g.small_category, c.verification_content))
                if isinstance(c.test_no, int):
                    ex_max[gk] = max(ex_max.get(gk, 0), c.test_no)

    counter = dict(ex_max)  # グループごとの現在No
    out_screens = []
    for sc in spec.screens:
        out_groups = []
        for g in sc.groups:
            gk = group_key(sc.screen_name, g.medium_category, g.small_category)
            out_cases = []
            for c in g.cases:
                if drop_dups and case_key(sc.screen_name, g.medium_category, g.small_category, c.verification_content) in ex_keys:
                    continue
                counter[gk] = counter.get(gk, 0) + 1
                out_cases.append(c.model_copy(update={"test_no": counter[gk], "source_row": None}))
            if out_cases:
                out_groups.append(Group(medium_category=g.medium_category, small_category=g.small_category, cases=out_cases))
        if out_groups:
            out_screens.append(Screen(screen_name=sc.screen_name, groups=out_groups))
    return TestSpec(sheet=spec.sheet, screens=out_screens)


def update_case(target: str, sheet: str, row: int, verification_content: str, steps: list[str], mapping=None) -> str:
    """既存ケース行（source_row）の確認内容・実施手順を上書きする（意味的マージの更新用）。"""
    from openpyxl.utils import get_column_letter

    from gwex.domains.model import DEFAULT_MAPPING
    from gwex.fetcher import google

    cfg = mapping or DEFAULT_MAPPING
    vcol = get_column_letter(cfg.column_index("verification_content") + 1)
    scol = get_column_letter(cfg.column_index("execution_steps") + 1)
    asgn = [(f"{vcol}{row}", verification_content), (f"{scol}{row}", "\n".join(steps))]
    if google.detect(target) is not None:
        from gwex.writer import gsheet_writer

        return gsheet_writer.write_cells(target, sheet, asgn)
    from gwex.writer import xlsx_writer

    return xlsx_writer.write_cells(target, sheet, asgn)


def load_mapping(path: str):
    """TOML から MappingConfig を読む（privileged: ファイル I/O）。"""
    import tomllib

    from gwex.domains.model import MappingConfig

    with open(path, "rb") as f:
        raw = tomllib.load(f)
    options = raw.pop("options", {})
    return MappingConfig(**raw, **options)
