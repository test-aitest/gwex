"""概要設計書テンプレの「記入し残し（プレースホルダ／サンプル／指示文）」を機械検出する。

完成宣言の前に必ず通す lint。人の記憶に頼らず、既知の残骸パターンを網羅走査して
セル番地つきで報告する（2026-000322 で書式番号 `2026-000xxx[…してください]` を
目視走査で取りこぼした反省）。

- ERROR: ほぼ確実に消し残し（案件番号・要素プレースホルダ・サンプル文言）
- WARN : 文脈次第（フォーム標準の指示文。残す場合もあるので人が判断）
"""

from __future__ import annotations

import re
from typing import Optional

from openpyxl import load_workbook

# 確実に残骸（消す/置換する）
ERROR_PATTERNS = [
    r"\{[A-Z]{1,3}\d",          # {C10} {AE6} {BI6:BL6} 等
    r"\[あなた",                 # [あなたが依頼内容…]
    r"〇〇", r"△△",             # サンプル記号
    r"000xxx",                  # 案件番号プレースホルダ（書式番号セル）
    r"\[No",                    # [Noより後の値…]
    r"_template",               # ファイル名残骸
    r"振込で借りる",             # テンプレ作成者のサンプル文言
]

# 文脈次第（フォーム標準ガイド。answer では消える場合もあるので人が確認）
WARN_PATTERNS = [
    r"してください",
    r"記入すること",
    r"添付してください",
    # 以下2つは案件が正規に記入する大項目名でもある（テンプレ指示が書かせる標準項目）。
    # サンプル残骸の場合は同行の {G16}[…] 等が ERROR_PATTERNS で検出されるため、ここは人の判断に回す
    r"APIレスポンスのマスキング",  # サンプルテスト項目 or 正規の大項目
    r"B→dashのイベント",          # サンプルテスト項目 or 正規の大項目
]

# 走査対象から除外する語（テスト後に記入する正規の空欄など）
DEFAULT_IGNORE = ["回答してください"]


def lint(path: str, *, sheets: Optional[list[str]] = None,
         extra_ignore: Optional[list[str]] = None) -> dict:
    """workbook を走査し残骸を返す。返り値 {errors:[…], warnings:[…]}（各 {sheet,cell,text,pattern}）。"""
    wb = load_workbook(path)
    targets = sheets or wb.sheetnames
    ignore = DEFAULT_IGNORE + list(extra_ignore or [])
    errors, warnings = [], []
    for s in targets:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v.strip():
                    continue
                if any(ig in v for ig in ignore):
                    continue
                hit = {"sheet": s, "cell": c.coordinate, "text": v[:60]}
                for pat in ERROR_PATTERNS:
                    if re.search(pat, v):
                        errors.append({**hit, "pattern": pat})
                        break
                else:
                    for pat in WARN_PATTERNS:
                        if re.search(pat, v):
                            warnings.append({**hit, "pattern": pat})
                            break
    return {"errors": errors, "warnings": warnings}


def summarize(result: dict) -> str:
    e, w = result["errors"], result["warnings"]
    lines = [f"=== lint: ERROR {len(e)} / WARN {len(w)} ==="]
    for h in e:
        lines.append(f"  [ERROR] {h['sheet']}!{h['cell']}: {h['text']!r}  (/{h['pattern']}/)")
    for h in w:
        lines.append(f"  [warn]  {h['sheet']}!{h['cell']}: {h['text']!r}  (/{h['pattern']}/)")
    if not e:
        lines.append("ERROR 0 → 残骸なし（WARN は文脈確認）。")
    return "\n".join(lines)
