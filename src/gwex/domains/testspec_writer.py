"""TestSpec（階層）→ セル割当 [(cell, value)] への展開。pure な配置計算。

抽出の逆変換。結合を作らずに「親（画面/中項目/小項目）は先頭行に1回だけ」書き、
継続行は空欄にする（抽出側がカテゴリを前方補完して復元する）。test_no/確認内容/
実施手順は各ケース行に書く（execution_steps は改行結合で 1 セル）。
"""

from __future__ import annotations

from typing import Optional

from openpyxl.utils import get_column_letter

from gwex.domains.model import DEFAULT_MAPPING, MappingConfig, TestSpec

_ROLES = (
    "screen_name",
    "medium_category",
    "small_category",
    "test_no",
    "verification_content",
    "execution_steps",
)


def assignments(
    spec: TestSpec,
    mapping: Optional[MappingConfig] = None,
    start_row: Optional[int] = None,
) -> list[tuple[str, object]]:
    cfg = mapping or DEFAULT_MAPPING
    col = {}
    for role in _ROLES:
        idx = cfg.column_index(role)
        if idx is not None:
            col[role] = get_column_letter(idx + 1)

    out: list[tuple[str, object]] = []
    row = start_row if start_row is not None else cfg.data_start_row

    def put(role: str, value: object) -> None:
        if role in col and value not in (None, ""):
            out.append((f"{col[role]}{row}", value))

    for screen in spec.screens:
        first_screen = True
        for group in screen.groups:
            first_group = True
            for case in group.cases:
                if first_screen:
                    put("screen_name", screen.screen_name)
                    first_screen = False
                if first_group:
                    put("medium_category", group.medium_category)
                    put("small_category", group.small_category)
                    first_group = False
                put("test_no", case.test_no)
                put("verification_content", case.verification_content)
                put("execution_steps", "\n".join(case.execution_steps))
                row += 1
    return out


def layout(
    spec: TestSpec,
    mapping: Optional[MappingConfig] = None,
    start_row: Optional[int] = None,
    screen_start_no: int = 1,
) -> dict:
    """整形用の配置情報（親セルの縦結合スパンと採番）を計算する。pure。

    返り値:
      data_first_row / data_last_row: 書き込んだケース行の範囲(1始まり)
      merges: 縦結合すべき範囲レター文字列 ["B113:B115", ...]（画面/中項目/小項目の No 列・値列）
      numbers: [(cell, no)] 親番号（画面No=連番, 中項目No=画面内連番, 小項目No=中項目内連番）

    画面は全ケース行、中項目は同一中項目の連続グループ、小項目は1グループでスパンを取る。
    番号列・結合は mapping.number_columns / columns に定義された列にのみ適用。
    """
    cfg = mapping or DEFAULT_MAPPING
    val = {
        r: get_column_letter(cfg.column_index(r) + 1)
        for r in ("screen_name", "medium_category", "small_category")
        if cfg.column_index(r) is not None
    }
    num = dict(cfg.number_columns or {})
    row = start_row if start_row is not None else cfg.data_start_row
    first = row
    merges: list[str] = []
    numbers: list[tuple[str, int]] = []

    def span(role: str, r0: int, r1: int, no: int) -> None:
        for table in (val, num):
            c = table.get(role)
            if c and r1 > r0:
                merges.append(f"{c}{r0}:{c}{r1}")
        if role in num:
            numbers.append((f"{num[role]}{r0}", no))

    sno = screen_start_no
    for screen in spec.screens:
        s0 = row
        mno = 0
        j = 0
        groups = screen.groups
        while j < len(groups):
            med = groups[j].medium_category
            k = j + 1
            while k < len(groups) and groups[k].medium_category == med:
                k += 1
            m0 = row
            mno += 1
            for gi in range(j, k):
                g0 = row
                row += len(groups[gi].cases)
                span("small_category", g0, row - 1, gi - j + 1)
            span("medium_category", m0, row - 1, mno)
            j = k
        span("screen_name", s0, row - 1, sno)
        sno += 1

    return {
        "data_first_row": first,
        "data_last_row": row - 1,
        "merges": merges,
        "numbers": numbers,
    }
