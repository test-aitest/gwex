"""2つの .xlsx を構造比較する（anken-verify 手順1「answer 全セル diff」用）。

値・結合・画像（枚数＋位置/寸法）・alignment・行高さの5観点を比較する。比較対象パスは
引数で受け、ハードコードしない（アドホックスクリプトの置き換え）。

画像は枚数だけでなく **アンカーセルと表示寸法（EMU）** まで比較する。枚数が同じでも
貼る位置や大きさが違えば見た目は別物になるため（枚数一致で「OK」と誤判定した実績あり）。
1枚ごとの破綻（はみ出し・浮遊・二重）は `gwex check-images` が担当する。
"""

from __future__ import annotations

from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _align(cell):
    a = cell.alignment
    return (a.horizontal, a.vertical, bool(a.wrap_text))


def _images_by_cell(ws) -> dict:
    """{アンカーセル: (幅EMU, 高さEMU)}。ext を持たないアンカーは無視する。"""
    from gwex.domains.image_check import _describe

    out = {}
    for im in ws._images:
        d = _describe(ws, im)
        if d:
            out[d["cell"]] = (d["w_emu"], d["h_emu"])
    return out


def _compare_images(wsa, wsb) -> list[dict]:
    """アンカーセル単位で画像の有無と表示寸法を突き合わせる。"""
    ia, ib = _images_by_cell(wsa), _images_by_cell(wsb)
    diffs = []
    for cell in sorted(set(ia) | set(ib)):
        a, b = ia.get(cell), ib.get(cell)
        if a == b:
            continue
        diffs.append({
            "cell": cell,
            "a": None if a is None else f"{round(a[0]/9525,1)}x{round(a[1]/9525,1)}px",
            "b": None if b is None else f"{round(b[0]/9525,1)}x{round(b[1]/9525,1)}px",
        })
    return diffs


def _compare_sheet(wsa, wsb, max_row: Optional[int], max_col: Optional[int]) -> dict:
    mr = max_row or max(wsa.max_row, wsb.max_row)
    mc = max_col or max(wsa.max_column, wsb.max_column)
    values, alignment = [], []
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            ca, cb = wsa.cell(r, c), wsb.cell(r, c)
            addr = f"{get_column_letter(c)}{r}"
            if ca.value != cb.value:
                values.append({"cell": addr, "a": _s(ca.value), "b": _s(cb.value)})
            if _align(ca) != _align(cb):
                alignment.append({"cell": addr, "a": _align(ca), "b": _align(cb)})
    sa = {str(m) for m in wsa.merged_cells.ranges}
    sb = {str(m) for m in wsb.merged_cells.ranges}
    heights = []
    for r in range(1, mr + 1):
        ha, hb = wsa.row_dimensions[r].height, wsb.row_dimensions[r].height
        if ha != hb:
            heights.append({"row": r, "a": ha, "b": hb})
    return {
        "values": values,
        "alignment": alignment,
        "merges_only_a": sorted(sa - sb),
        "merges_only_b": sorted(sb - sa),
        "images": {"a": len(wsa._images), "b": len(wsb._images)},
        "image_placements": _compare_images(wsa, wsb),
        "row_heights": heights,
    }


def _s(v):
    return str(v) if v is not None else None


def compare(a_path: str, b_path: str, *, sheet: Optional[str] = None,
            max_row: Optional[int] = None, max_col: Optional[int] = None) -> dict:
    """a_path（自作）と b_path（answer 等）を比較。sheet 省略時は共通シート全て。"""
    wba = load_workbook(a_path)
    wbb = load_workbook(b_path)
    sheets = [sheet] if sheet else [s for s in wba.sheetnames if s in wbb.sheetnames]
    out = {}
    for s in sheets:
        out[s] = _compare_sheet(wba[s], wbb[s], max_row, max_col)
    return out


def summarize(result: dict) -> str:
    """compare() の結果を人間可読サマリ（観点別件数）に整形する。"""
    lines = []
    total = 0
    for s, d in result.items():
        placements = d.get("image_placements", [])
        n = (len(d["values"]) + len(d["alignment"]) + len(d["merges_only_a"])
             + len(d["merges_only_b"]) + len(d["row_heights"]) + len(placements)
             + (1 if d["images"]["a"] != d["images"]["b"] else 0))
        total += n
        lines.append(
            f"[{s}] 値:{len(d['values'])} 整列:{len(d['alignment'])} "
            f"結合(自作のみ:{len(d['merges_only_a'])}/正解のみ:{len(d['merges_only_b'])}) "
            f"画像:{d['images']['a']}vs{d['images']['b']}(配置差:{len(placements)}) "
            f"行高:{len(d['row_heights'])}"
        )
        for v in d["values"][:20]:
            lines.append(f"    値DIFF {v['cell']}: A={v['a']!r} B={v['b']!r}")
        for p in placements[:20]:
            lines.append(f"    画像DIFF {p['cell']}: A={p['a']} B={p['b']}")
    lines.append(f"=== 総差分件数: {total}（0 なら一致）===")
    return "\n".join(lines)
