"""シートに配置された画像の「見た目の破綻」を機械検出する。

`gwex lint`（文字列の残骸）と `gwex diff`（値・整列・結合・画像枚数・行高）が
見ていない領域＝**画像の位置と寸法**を数値で検査する。人が PNG を目視して
「はみ出している気がする」「前後でサイズが違う気がする」と判定するのを置き換える。

検出する破綻（概要設計書「2.画面イメージ」で実際に起きたもの）:

- ``overflow``      画像の表示寸法が、アンカー先の結合枠より大きい（枠からはみ出す）
- ``no_frame``      画像が結合枠の左上に紐づいていない（セルの上に浮いている）
- ``pair_mismatch`` 修正前／修正後のペアで表示寸法が違う
- ``duplicate``     同一アンカーセルに画像が2枚以上ある（埋め込み＋overlay の二重貼り）
- ``aspect``        表示寸法の縦横比が元画像とずれている（引き伸ばし／潰れ）

寸法は OOXML の EMU（English Metric Unit, 9525 EMU = 1 px）で保持されるため、
``pair_mismatch`` と ``duplicate`` は**誤差ゼロの厳密判定**ができる。
``overflow`` だけは結合枠の実寸を列幅（文字数単位）・行高（pt）から px へ換算する
必要があり、フォント依存の誤差が乗るので ``tolerance_px`` で吸収する。
"""

from __future__ import annotations

from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

EMU_PER_PX = 9525

# 列幅（文字数）→ px。Excel の標準フォント基準の慣例式。
_COL_WIDTH_DEFAULT = 8.43
# 行高（pt）→ px。96dpi / 72pt。
_ROW_HEIGHT_DEFAULT = 15.0
_PT_TO_PX = 96.0 / 72.0


def _col_px(ws, col_idx: int) -> float:
    d = ws.column_dimensions.get(get_column_letter(col_idx))
    w = getattr(d, "width", None) if d is not None else None
    if w is None:
        w = ws.sheet_format.defaultColWidth or _COL_WIDTH_DEFAULT
    return round(w * 7) + 5


def _row_px(ws, row_idx: int) -> float:
    d = ws.row_dimensions.get(row_idx)
    h = getattr(d, "height", None) if d is not None else None
    if h is None:
        h = ws.sheet_format.defaultRowHeight or _ROW_HEIGHT_DEFAULT
    return h * _PT_TO_PX


def _frame_for(ws, anchor_cell: str) -> Optional[str]:
    """anchor_cell を左上に持つ結合範囲を返す（無ければ None）。"""
    for m in ws.merged_cells.ranges:
        if m.coord.split(":")[0] == anchor_cell:
            return m.coord
    return None


def _frame_size_px(ws, coord: str) -> tuple[float, float]:
    min_c, min_r, max_c, max_r = range_boundaries(coord)
    w = sum(_col_px(ws, c) for c in range(min_c, max_c + 1))
    h = sum(_row_px(ws, r) for r in range(min_r, max_r + 1))
    return w, h


def _describe(ws, img) -> Optional[dict]:
    a = img.anchor
    frm = getattr(a, "_from", None)
    ext = getattr(a, "ext", None)
    if frm is None or ext is None:
        # TwoCellAnchor 等、ext を持たないアンカーは寸法比較の対象外
        return None
    cell = f"{get_column_letter(frm.col + 1)}{frm.row + 1}"
    frame = _frame_for(ws, cell)
    d = {
        "cell": cell,
        "row": frm.row + 1,
        "col": frm.col + 1,
        "w_emu": ext.cx,
        "h_emu": ext.cy,
        "w_px": round(ext.cx / EMU_PER_PX, 1),
        "h_px": round(ext.cy / EMU_PER_PX, 1),
        "frame": frame,
        # img.width/height は「元画像のピクセル寸法」（表示寸法ではない）。
        # 表示との縦横比のずれ＝引き伸ばし／潰れの検出に使う。
        "src_w": getattr(img, "width", None),
        "src_h": getattr(img, "height", None),
    }
    if frame:
        fw, fh = _frame_size_px(ws, frame)
        d["frame_w_px"], d["frame_h_px"] = round(fw, 1), round(fh, 1)
    return d


def check(path: str, *, sheets: Optional[list[str]] = None,
          tolerance_px: float = 4.0, aspect_tolerance: float = 0.01) -> dict:
    """画像の位置・寸法を検査する。返り値 {images: {...}, issues: [...]}。

    issues の各要素は {kind, sheet, cell, detail} 形式。空なら破綻なし。
    aspect_tolerance は縦横比のずれの許容割合（既定 1%）。
    """
    wb = load_workbook(path)
    targets = sheets or wb.sheetnames
    all_images: dict[str, list[dict]] = {}
    issues: list[dict] = []

    for s in targets:
        if s not in wb.sheetnames:
            continue
        ws = wb[s]
        descs = [d for d in (_describe(ws, im) for im in ws._images) if d]
        all_images[s] = descs

        seen: dict[str, int] = {}
        for d in descs:
            seen[d["cell"]] = seen.get(d["cell"], 0) + 1

            if d["src_w"] and d["src_h"] and d["h_px"]:
                src_ratio = d["src_w"] / d["src_h"]
                disp_ratio = d["w_px"] / d["h_px"]
                if abs(disp_ratio - src_ratio) / src_ratio > aspect_tolerance:
                    issues.append({
                        "kind": "aspect", "sheet": s, "cell": d["cell"],
                        "detail": f"縦横比が元画像とずれている（引き伸ばし/潰れ）: "
                                  f"元 {d['src_w']}x{d['src_h']}(比 {src_ratio:.4f}) → "
                                  f"表示 {d['w_px']}x{d['h_px']}(比 {disp_ratio:.4f})",
                    })

            if d["frame"] is None:
                issues.append({
                    "kind": "no_frame", "sheet": s, "cell": d["cell"],
                    "detail": f"結合枠の左上に紐づかない画像（{d['w_px']}x{d['h_px']}px）。"
                              "セルの上に浮いている可能性",
                })
                continue

            over_w = d["w_px"] - d["frame_w_px"]
            over_h = d["h_px"] - d["frame_h_px"]
            if over_w > tolerance_px or over_h > tolerance_px:
                issues.append({
                    "kind": "overflow", "sheet": s, "cell": d["cell"],
                    "detail": f"画像 {d['w_px']}x{d['h_px']}px が枠 {d['frame']} "
                              f"({d['frame_w_px']}x{d['frame_h_px']}px) をはみ出す "
                              f"(+{round(max(over_w, 0), 1)}w/+{round(max(over_h, 0), 1)}h)",
                })

        for cell, n in seen.items():
            if n > 1:
                issues.append({
                    "kind": "duplicate", "sheet": s, "cell": cell,
                    "detail": f"同一アンカーに画像が {n} 枚（埋め込み＋overlay の二重貼り）",
                })

        # 修正前／修正後のペアは表示寸法が一致すべき。
        # 1行に複数ペアが並ぶ（テンプレは左ペア C/H と右ペア N/S の2組）ため、
        # 「同じ行の画像すべてが同寸」ではなく、列順に2枚ずつ組にして比較する。
        by_row: dict[int, list[dict]] = {}
        for d in descs:
            by_row.setdefault(d["row"], []).append(d)
        for row, group in sorted(by_row.items()):
            ordered = sorted(group, key=lambda d: d["col"])
            for i in range(0, len(ordered) - 1, 2):
                before, after = ordered[i], ordered[i + 1]
                if (before["w_emu"], before["h_emu"]) != (after["w_emu"], after["h_emu"]):
                    issues.append({
                        "kind": "pair_mismatch", "sheet": s, "cell": before["cell"],
                        "detail": f"修正前/修正後の表示寸法が不一致: "
                                  f"{before['cell']}={before['w_px']}x{before['h_px']}px / "
                                  f"{after['cell']}={after['w_px']}x{after['h_px']}px",
                    })

    return {"images": all_images, "issues": issues}


def summarize(result: dict) -> str:
    issues = result["issues"]
    n_img = sum(len(v) for v in result["images"].values())
    lines = [f"=== check-images: 画像 {n_img} 枚 / 破綻 {len(issues)} 件 ==="]
    for s, descs in result["images"].items():
        if not descs:
            continue
        lines.append(f"[{s}] {len(descs)} 枚")
        for d in descs:
            frame = d["frame"] or "(枠なし)"
            fit = ""
            if d["frame"]:
                fit = f" ⊂ {frame}({d['frame_w_px']}x{d['frame_h_px']})"
            lines.append(f"    {d['cell']}: {d['w_px']}x{d['h_px']}px{fit}")
    for i in issues:
        lines.append(f"  [NG:{i['kind']}] {i['sheet']}!{i['cell']}: {i['detail']}")
    if not issues:
        lines.append("破綻 0 → はみ出し・浮遊・二重・前後サイズ違い なし。")
    return "\n".join(lines)
