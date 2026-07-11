"""flow_graph: スクリーングラフ → draw-flow spec 雛形の生成。

経路探索（BFS・via 経由・エラー）、ラベル解決（component label / 名前フォールバック）、
YAML 雛形のラウンドトリップ、draw-flow までの統合を検証する。
"""

from __future__ import annotations

import pytest
import yaml

from gwex.domains import flow_graph

GRAPH = {
    "A": {
        "components": [
            {"name": "btn1", "type": "BUTTON", "accessibilityId": "btn1", "label": "進む"},
        ],
        "navigators": [
            {"targetScreen": "B", "type": "TAP", "triggerComponent": "btn1",
             "requiredInputs": {"loginid": "SECRET-ID", "loginpass": "SECRET-PW"}},
        ],
    },
    "B": {
        "components": [
            {"name": "btn2", "type": "BUTTON", "accessibilityId": "btn2", "label": ""},
        ],
        "navigators": [
            {"targetScreen": "C", "type": "TAP", "triggerComponent": "btn2"},
            {"targetScreen": "A", "type": "TAP", "triggerComponent": "back"},
        ],
    },
    "C": {"components": [], "navigators": []},
}


def test_find_path_bfs_and_via():
    assert flow_graph.find_path(GRAPH, "A", "C") == ["A", "B", "C"]
    assert flow_graph.find_path(GRAPH, "A", "C", via=["B"]) == ["A", "B", "C"]
    assert flow_graph.find_path(GRAPH, "A", "A") == ["A"]


def test_find_path_errors():
    with pytest.raises(ValueError, match="画面がグラフにありません"):
        flow_graph.find_path(GRAPH, "A", "Z")
    with pytest.raises(ValueError, match="経路が見つかりません"):
        flow_graph.find_path(GRAPH, "C", "A")  # C に navigator が無い


def test_build_spec_labels_and_layout():
    spec = flow_graph.build_spec(GRAPH, ["A", "B", "C"])
    assert [(n["id"], n["col"], n["row"]) for n in spec["nodes"]] == [
        ("A", 0, 0), ("B", 1, 0), ("C", 2, 0)]
    assert spec["nodes"][0]["image"] == "caps/A.png"
    e1, e2 = spec["edges"]
    assert (e1["from"], e1["to"], e1["label"]) == ("A", "B", "進む")   # component label
    assert (e2["label"], e2["trigger_component"]) == ("btn2", "btn2")  # label 空 → 名前で代用
    assert len(spec["sheet"]) <= 31  # Excel シート名上限


def test_render_yaml_roundtrip_no_secrets():
    spec = flow_graph.build_spec(GRAPH, ["A", "B", "C"])
    text = flow_graph.render_spec_yaml(spec)
    assert "TODO" in text and "trigger_rect" in text
    assert "SECRET" not in text  # requiredInputs（認証情報）を転記しない
    loaded = yaml.safe_load(text)
    assert [n["id"] for n in loaded["nodes"]] == ["A", "B", "C"]
    assert loaded["edges"][0]["label"] == "進む"


def test_integration_with_draw_flow(tmp_path):
    """flow-spec の出力（画像だけ差し替え）がそのまま draw-flow に通る。"""
    PIL = pytest.importorskip("PIL.Image")
    from gwex.writer import xlsx_flow

    (tmp_path / "caps").mkdir()
    for s in ["A", "B", "C"]:
        PIL.new("RGB", (100, 200), (60, 60, 60)).save(tmp_path / "caps" / f"{s}.png")

    graph_yaml = tmp_path / "graph.yaml"
    graph_yaml.write_text(yaml.safe_dump({"screens": GRAPH}, allow_unicode=True), encoding="utf-8")
    screens = flow_graph.load_graph(str(graph_yaml))
    spec = yaml.safe_load(flow_graph.render_spec_yaml(
        flow_graph.build_spec(screens, flow_graph.find_path(screens, "A", "C"))))

    out = str(tmp_path / "flow.xlsx")
    xlsx_flow.draw_flow(spec, out, base_dir=str(tmp_path))
    from openpyxl import load_workbook
    assert load_workbook(out).sheetnames == [spec["sheet"]]


def test_load_graph_requires_screens(tmp_path):
    p = tmp_path / "bad.yaml"
    p.write_text("foo: 1", encoding="utf-8")
    with pytest.raises(ValueError, match="screens"):
        flow_graph.load_graph(str(p))
