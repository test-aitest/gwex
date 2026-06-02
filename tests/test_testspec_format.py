"""testspec 追記ブロックの体裁整形（罫線/縦結合/親番号）の単体テスト。"""

from __future__ import annotations

from io import BytesIO

import openpyxl

from gwex.domains.model import DEFAULT_MAPPING
from gwex.domains import testspec_writer


def test_layout_merges_and_numbers():
    """単一画面/中項目/小項目・3ケースで親が縦結合され、画面Noが継続採番されること。"""
    from gwex.domains.model import Case, Group, Screen, TestSpec
    spec = TestSpec(sheet="S", screens=[Screen(screen_name="ダッシュボード", groups=[
        Group(medium_category="金額表示", small_category="前回ご返済額", cases=[
            Case(test_no=1, verification_content="v1"),
            Case(test_no=2, verification_content="v2"),
            Case(test_no=3, verification_content="v3"),
        ])])])
    plan = testspec_writer.layout(spec, DEFAULT_MAPPING, start_row=113, screen_start_no=14)
    assert plan["data_first_row"] == 113 and plan["data_last_row"] == 115
    # 親（No列B/D/F・値列C/E/G）が 113:115 で縦結合
    for ref in ["B113:B115", "C113:C115", "D113:D115", "E113:E115", "F113:F115", "G113:G115"]:
        assert ref in plan["merges"], ref
    # 番号: 画面=14（継続）, 中項目=1, 小項目=1
    nums = dict(plan["numbers"])
    assert nums["B113"] == 14 and nums["D113"] == 1 and nums["F113"] == 1


def test_format_applies_borders_to_box_right(tmp_path):
    """core.write_testspec(apply_format=True) で追記行が箱罫線(B..S)＋結合＋番号になること。"""
    from gwex import core
    # 既存テンプレ風シート: ヘッダ行7 + 既存1画面(B8=1)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.結合テスト項目"
    ws["C7"], ws["E7"], ws["G7"] = "画面名", "中項目名", "小項目名"
    ws["H7"], ws["I7"], ws["J7"] = "No", "確認内容", "実施手順"
    ws["B10"], ws["C10"], ws["E10"], ws["G10"] = 1, "既存画面", "中", "小"
    ws["H10"], ws["I10"] = 1, "既存確認"
    p = tmp_path / "t.xlsx"
    wb.save(p)

    spec_json = (
        '{"sheet":"3.結合テスト項目","screens":[{"screen_name":"ダッシュボード",'
        '"groups":[{"medium_category":"金額表示","small_category":"前回ご返済額",'
        '"cases":[{"test_no":1,"verification_content":"a","execution_steps":["①x"]},'
        '{"test_no":2,"verification_content":"b","execution_steps":["①y"]}]}]}]}'
    )
    core.write_testspec(spec_json, str(p), "3.結合テスト項目", None,
                        append=True, dedup=True, apply_format=True)

    ws2 = openpyxl.load_workbook(p)["3.結合テスト項目"]
    # 追記行（既存last=10 の下 → 11..12）
    r0 = 11
    # 罫線が S(19) まで
    def right_border_col(r):
        last = 0
        for c in range(2, 25):
            b = ws2.cell(row=r, column=c).border
            if any(sd and sd.style for sd in [b.left, b.right, b.top, b.bottom]):
                last = c
        return last
    assert right_border_col(r0) == 19  # S
    assert right_border_col(r0 + 1) == 19
    # 親縦結合（2ケース）
    merged = [str(m) for m in ws2.merged_cells.ranges]
    assert f"C{r0}:C{r0+1}" in merged and f"B{r0}:B{r0+1}" in merged
    # 画面No は既存(1)の次=2
    assert ws2.cell(row=r0, column=2).value == 2


def test_create_before_after_section(tmp_path):
    """枠付き before/after セクション: 見出し・ラベル結合・箱罫線・画像配置。"""
    import pytest
    PIL = pytest.importorskip("PIL.Image")
    from gwex.writer import xlsx_format

    # 縦長スクショ（高さが幅の倍）→ 必要行数が複数になる
    shot_b = tmp_path / "b.png"; PIL.new("RGB", (300, 650), (0, 128, 255)).save(shot_b)
    shot_a = tmp_path / "a.png"; PIL.new("RGB", (300, 650), (200, 0, 40)).save(shot_a)
    p = tmp_path / "img.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "2.画面イメージ(iOS)"; wb.save(p)

    xlsx_format.create_before_after_section(
        str(p), "2.画面イメージ(iOS)", 100, "ダッシュボード",
        str(shot_b), str(shot_a), left_cols=("C", "L"), split_col="H",
    )
    ws = openpyxl.load_workbook(p)["2.画面イメージ(iOS)"]
    merged = [str(m) for m in ws.merged_cells.ranges]
    assert "C100:L100" in merged          # 見出し
    assert "C101:G101" in merged          # 修正前ラベル
    assert "H101:L101" in merged          # 修正後ラベル
    assert ws["C100"].value == "ダッシュボード"
    assert ws["C101"].value == "修正前" and ws["H101"].value == "修正後"
    # 見出し・ラベルは中央揃え
    assert ws["C100"].alignment.horizontal == "center"
    assert ws["C101"].alignment.horizontal == "center" and ws["H101"].alignment.horizontal == "center"
    assert ws["C100"].border.top.style == "thin"
    # 画像エリアが結合されている（修正前 C..G / 修正後 H..L、行102から）
    img_merges = [str(m) for m in ws.merged_cells.ranges if m.min_row >= 102]
    assert any(m.startswith("C102:G") for m in img_merges), img_merges
    assert any(m.startswith("H102:L") for m in img_merges), img_merges
    assert len(ws._images) == 2           # 前/後（xlsx_zip 配置）
    # 画像は枠の80%サイズ（ぴったりではない）・巨大化していない
    widths = sorted(im.width for im in ws._images)
    assert all(w < 600 for w in widths), widths
    # 背景青（見出し=濃い青、画像エリア=薄い青）。画像エリアは結合済みなので先頭C102を見る
    assert ws["C100"].fill.fgColor.rgb.endswith("9CC2E5")   # 見出し
    assert ws["C102"].fill.fgColor.rgb.endswith("DEEAF6")   # 画像エリア(結合先頭)
    # 比率維持: ext 比 ≈ 元画像比(300/650)
    for im in ws._images:
        assert abs(im.width / im.height - 300 / 650) < 0.02
