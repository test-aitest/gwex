"""writer.xlsm_ops.apply_ops: 宣言的 ops の ZIP 手術適用。

検証の柱:
- 最小タッチ（変更された ZIP エントリ集合の assert・vba/sharedStrings 不変）
- insert_rows 後の下方 図形アンカー・結合セルのシフト
- edit_shape_text の NFC マッチ（NFD 入力でも一致）と 0件/複数件の明示エラー
- 失敗時は何も書かない（原子性）
"""

from __future__ import annotations

import datetime
import re
import unicodedata
import zipfile

import pytest
from openpyxl import load_workbook

from gwex.writer import xlsm_ops

from tests._xlsm_fixtures import entries_of, make_book, rewrite

PIL = pytest.importorskip("PIL.Image")


def _apply(p, tmp_path, ops, sheet="N1"):
    out = str(tmp_path / "out.xlsx")
    report = xlsm_ops.apply_ops(p, {"sheet": sheet, "ops": ops}, output=out)
    return out, report


_ROW11 = '<row r="11"><c r="B11" s="7"><v>1</v></c><c r="D11" s="8"><v>2</v></c></row>'


def _inject_self_closing_row(p):
    """sheet1 に実ファイル形状の自己終了行 r=10 と、直後のセル入り行 r=11 を注入する。"""
    entries = entries_of(p)
    sx = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert '<row r="12"' in sx
    sx = sx.replace('<row r="12"',
                    '<row r="10" ht="20" customHeight="1"/>' + _ROW11 + '<row r="12"', 1)
    entries["xl/worksheets/sheet1.xml"] = sx.encode("utf-8")
    rewrite(p, entries)


def _changed(p, out):
    before, after = entries_of(p), entries_of(out)
    changed = sorted(n for n in before if n in after and before[n] != after[n])
    added = sorted(n for n in after if n not in before)
    return changed, added


def test_set_cell_minimal_touch(tmp_path):
    """変更されるのは対象シート XML ただ1つ。vba・sharedStrings はバイト不変。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"set_cell": {"cell": "B8", "text": "新しい説明", "mode": "replace"}},
    ])
    assert report["ok"], report
    changed, added = _changed(p, out)
    assert changed == ["xl/worksheets/sheet1.xml"]
    assert added == []
    assert report["changed_entries"] == ["xl/worksheets/sheet1.xml"]

    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    m = re.search(r'<c r="B8"[^>]*t="inlineStr"><is><t[^>]*>([^<]*)</t></is></c>', sx)
    assert m and m.group(1) == "新しい説明"


def test_set_cell_append_joins_with_newline(tmp_path):
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"set_cell": {"cell": "B8", "text": "追記行", "mode": "append"}},
    ])
    assert report["ok"], report
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    m = re.search(r'<c r="B8"[^>]*t="inlineStr"><is><t[^>]*>(.*?)</t></is></c>', sx, re.DOTALL)
    assert m.group(1) == "ダッシュボード概要\n追記行"


def test_set_cell_into_self_closing_row_expands_without_touching_next_row(tmp_path):
    """自己終了 <row r="10"/> への書き込みは行を展開し、次行 r=11 を壊さない。

    旧regexは <row .../> の次の </row> まで飲み込み、正しい r 属性のまま
    B10 を row 11 の中に列順を壊して注入していた（過去実害の破損クラス）。
    """
    p = make_book(tmp_path)
    _inject_self_closing_row(p)
    out, report = _apply(p, tmp_path, [
        {"set_cell": {"cell": "B10", "text": "展開された行"}},
    ])
    assert report["ok"], report
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert re.search(
        r'<row r="10" ht="20" customHeight="1"><c r="B10" t="inlineStr">'
        r"<is><t[^>]*>展開された行</t></is></c></row>", sx)
    assert _ROW11 in sx  # 次行はバイト同一のまま


def test_set_cell_insert_keeps_column_order(tmp_path):
    """既存セルがある行への挿入は列順（A,B,…）を保つ（先頭・間・末尾）。"""
    p = make_book(tmp_path)
    _inject_self_closing_row(p)
    out, report = _apply(p, tmp_path, [
        {"set_cell": {"cell": "C11", "text": "間に挿入"}},
        {"set_cell": {"cell": "A11", "number": 1}},
        {"set_cell": {"cell": "E11", "number": 2}},
    ])
    assert report["ok"], report
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    row11 = re.search(r'<row r="11">.*?</row>', sx).group(0)
    assert re.findall(r'<c r="([A-Z]+11)"', row11) == ["A11", "B11", "C11", "D11", "E11"]


def test_set_cell_number(tmp_path):
    p = make_book(tmp_path)
    _inject_self_closing_row(p)
    out, report = _apply(p, tmp_path, [
        {"set_cell": {"cell": "B11", "number": 46071}},
        {"set_cell": {"cell": "F11", "number": 0.5}},
    ])
    assert report["ok"], report
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert '<c r="B11" s="7"><v>46071</v></c>' in sx  # t なし＝数値・既存 s= 温存
    assert '<c r="F11"><v>0.5</v></c>' in sx

    for bad_params in ({"cell": "B11", "text": "x", "number": 1}, {"cell": "B11"}):
        bad = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"set_cell": bad_params}]},
                                 output=str(tmp_path / "x.xlsx"))
        assert not bad["ok"] and "どちらか一方" in bad["error"]
    assert not (tmp_path / "x.xlsx").exists()


def test_clear_cell_keeps_style_and_is_noop_when_missing(tmp_path):
    p = make_book(tmp_path)
    _inject_self_closing_row(p)
    out, report = _apply(p, tmp_path, [
        {"clear_cell": {"cell": "B11"}},   # s="7" のセル → 値だけ除去
        {"clear_cell": {"cell": "B8"}},    # スタイル無しの inlineStr セル
        {"clear_cell": {"cell": "Z99"}},   # 存在しないセル → no-op で成功
    ])
    assert report["ok"], report
    assert any("no-op" in s for s in report["ops"])
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert '<c r="B11" s="7"/>' in sx
    assert '<c r="D11" s="8"><v>2</v></c>' in sx  # 隣のセルは不変
    assert re.search(r'<c r="B8"[^>]*/>', sx)
    assert "ダッシュボード概要" not in sx


def test_insert_rows_shifts_merges_and_anchors(tmp_path):
    """挿入位置より下の結合セルと drawing アンカーだけが下へずれる。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"insert_rows": {"at": 10, "count": 2, "copy_style_from": 12}},
    ])
    assert report["ok"], report

    entries = entries_of(out)
    sx = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    merges = re.findall(r'<mergeCell ref="([^"]+)"/>', sx)
    assert "B8:D9" in merges       # 挿入位置より上は不変
    assert "B24:C25" in merges     # B22:C23 が +2
    assert "B22:C23" not in merges
    assert re.search(r'<c r="B14"[^>]*t="s"', sx)  # B12(■イベント) が 14 へ
    assert re.search(r'<row r="10"[^>]*>', sx)     # 複製スタイルの空行が挿入された

    dx = entries["xl/drawings/drawing1.xml"].decode("utf-8")
    # pic の oneCellAnchor は row0=19 → 21 へ（0始まり）
    assert "<xdr:row>21</xdr:row>" in dx
    assert "<xdr:row>19</xdr:row>" not in dx

    changed, added = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml", "xl/worksheets/sheet1.xml"]
    assert added == []


def test_append_history_row(tmp_path):
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"append_history_row": {"sheet": "改版履歴", "date": "2026-07-11",
                                "target": "N1", "status": "変更", "content": "文言修正"}},
    ], sheet=None)
    assert report["ok"], report
    serial = (datetime.date(2026, 7, 11) - datetime.date(1899, 12, 30)).days

    sx = entries_of(out)["xl/worksheets/sheet2.xml"].decode("utf-8")
    assert re.search(r'<c r="A6"[^>]*><v>%d</v></c>' % serial, sx)
    for ref, text in (("B6", "N1"), ("C6", "変更"), ("E6", "文言修正")):
        m = re.search(r'<c r="%s"[^>]*t="inlineStr"><is><t[^>]*>([^<]*)</t></is></c>' % ref, sx)
        assert m and m.group(1) == text, ref

    changed, _ = _changed(p, out)
    assert changed == ["xl/worksheets/sheet2.xml"]


def test_edit_shape_text_nfc_match(tmp_path):
    """NFD で渡された match_text でも NFC 正規化で一致する。"""
    p = make_book(tmp_path)
    nfd = unicodedata.normalize("NFD", "ボタン変更前")
    assert nfd != "ボタン変更前"  # 前提: 本当に分解されている
    out, report = _apply(p, tmp_path, [
        {"edit_shape_text": {"match_text": nfd, "new_text": "ボタン変更後"}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    assert "<a:t>ボタン変更後</a:t>" in dx
    assert "ボタン変更前" not in dx
    # rPr（書式）は元の run から継承される
    assert re.search(r'<a:r><a:rPr lang="ja-JP" sz="1600" b="1"/><a:t>ボタン変更後</a:t></a:r>', dx)

    changed, _ = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml"]


def test_edit_shape_text_zero_or_ambiguous_fails_without_write(tmp_path):
    p = make_book(tmp_path)
    out = str(tmp_path / "out.xlsx")

    report = xlsm_ops.apply_ops(
        p, {"sheet": "N1", "ops": [{"edit_shape_text": {"match_text": "存在しない文言", "new_text": "x"}}]},
        output=out)
    assert not report["ok"]
    assert "一致する図形がありません" in report["error"]
    assert not (tmp_path / "out.xlsx").exists()

    report = xlsm_ops.apply_ops(
        p, {"sheet": "N1", "ops": [{"edit_shape_text": {"match_text": "重複文言", "new_text": "x"}}]},
        output=out)
    assert not report["ok"]
    assert "2 個" in report["error"]
    assert not (tmp_path / "out.xlsx").exists()


def test_failure_mid_ops_writes_nothing(tmp_path):
    """途中の op が失敗したら、先行 op が成功していても何も書かない。"""
    p = make_book(tmp_path)
    out = str(tmp_path / "out.xlsx")
    report = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [
        {"set_cell": {"cell": "B8", "text": "先行成功"}},
        {"delete_shape": {"match_text": "存在しない"}},
    ]}, output=out)
    assert not report["ok"]
    assert report["failed_op"] == {"index": 2, "op": "delete_shape"}
    assert not (tmp_path / "out.xlsx").exists()


def test_add_and_delete_shape(tmp_path):
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"add_shape": {"preset": "wedgeRectCallout", "text": "新ラベル",
                       "at_px": [120, 340], "size_px": [160, 40],
                       "style_from_text": "ラベルA"}},
        {"delete_shape": {"match_text": "ボタン変更前"}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    assert "<a:t>新ラベル</a:t>" in dx
    assert "ボタン変更前" not in dx
    # 位置は px→EMU（×9525）。プリセットは複製元と違うので avLst 初期化
    new_sp = re.search(r"<xdr:absoluteAnchor>.*?</xdr:absoluteAnchor>", dx, re.DOTALL).group(0)
    assert f'<xdr:pos x="{120 * 9525}" y="{340 * 9525}"/>' in new_sp
    assert f'<a:ext cx="{160 * 9525}" cy="{40 * 9525}"/>' in new_sp
    assert '<a:prstGeom prst="wedgeRectCallout"><a:avLst/></a:prstGeom>' in new_sp
    # cNvPr id は重複しない
    ids = re.findall(r'<xdr:cNvPr [^>]*?id="(\d+)"', dx)
    assert len(ids) == len(set(ids))

    changed, added = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml"]
    assert added == []


def test_add_connector_between_shapes(tmp_path):
    """shape_text 両端指定: stCxn/endCxn 接続 + 実例流儀の xfrm（右辺中点→左辺中点）。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"add_connector": {"preset": "straightConnector1",
                           "from": {"shape_text": "ボタン変更前"},
                           "to": {"shape_text": "ラベルA"}}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    cxn = re.search(r"<xdr:absoluteAnchor>.*?</xdr:absoluteAnchor>", dx, re.DOTALL).group(0)
    assert '<a:stCxn id="101" idx="3"/>' in cxn   # 101 の右辺
    assert '<a:endCxn id="102" idx="1"/>' in cxn  # 102 の左辺
    # 幾何: 101 右辺中点 (1548640, 2182880) → 102 左辺中点 (2500000, 2182880)
    assert '<a:off x="1548640" y="2182880"/>' in cxn
    assert '<a:ext cx="951360" cy="0"/>' in cxn
    assert "flipH" not in cxn and "flipV" not in cxn
    assert '<a:ln w="19050">' in cxn              # 既定 line_pt=1.5
    assert '<a:tailEnd type="triangle"/>' in cxn
    assert '<a:solidFill><a:schemeClr val="tx1"/></a:solidFill>' in cxn
    ids = re.findall(r'<xdr:cNvPr [^>]*?id="(\d+)"', dx)
    assert len(ids) == len(set(ids))              # cNvPr id の一意採番
    changed, added = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml"]
    assert added == []


def test_add_connector_px_to_pic(tmp_path):
    """px 始点は接続なし、pic_index 終点は endCxn 接続。line_pt 反映と adj1 流儀。"""
    p = make_book(tmp_path, with_connectors=True)  # pic に a:xfrm キャッシュあり
    out, report = _apply(p, tmp_path, [
        {"add_connector": {"preset": "bentConnector3", "line_pt": 1.0,
                           "from": {"px": [20, 870]},
                           "to": {"pic_index": 1}}},
        {"add_connector": {"from": {"px": [400, 870]},
                           "to": {"pic_index": 1}}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    cxn, cxn2 = re.findall(r"<xdr:absoluteAnchor>.*?</xdr:absoluteAnchor>", dx, re.DOTALL)
    assert "<a:stCxn" not in cxn                  # px 端は接続なし
    assert re.search(r'<a:endCxn id="\d+" idx="1"/>', cxn)  # pic 左辺へ接続
    # 始点 px(20,870)→EMU(190500, 8286750) / 終点 pic 左辺中点 (1200000, 9238250)
    assert '<a:off x="190500" y="8286750"/>' in cxn
    assert '<a:ext cx="1009500" cy="951500"/>' in cxn
    assert "flipH" not in cxn and "flipV" not in cxn
    assert '<a:ln w="12700">' in cxn              # line_pt: 1.0
    assert '<a:avLst><a:gd name="adj1" fmla="val 50000"/></a:avLst>' in cxn
    # 2本目: 始点が pic の右側 → pic 右辺(idx=3)へ、終点が始点より左 → flipH
    assert re.search(r'<a:endCxn id="\d+" idx="3"/>', cxn2)
    assert 'flipH="1"' in cxn2 and "flipV" not in cxn2
    assert '<a:off x="2343000" y="8286750"/>' in cxn2   # pic 右辺 x=1200000+1143000
    assert '<a:ext cx="1467000" cy="951500"/>' in cxn2


def test_add_connector_errors_write_nothing(tmp_path):
    p = make_book(tmp_path)
    out = str(tmp_path / "out.xlsx")

    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"add_connector": {
        "from": {"shape_text": "重複文言"}, "to": {"shape_text": "ラベルA"}}}]}, output=out)
    assert not r["ok"] and "2 個" in r["error"]

    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"add_connector": {
        "preset": "circle", "from": {"px": [0, 0]}, "to": {"px": [1, 1]}}}]}, output=out)
    assert not r["ok"] and "preset" in r["error"]

    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"add_connector": {
        "from": {"px": [5, 5]}, "to": {"px": [5, 5]}}}]}, output=out)
    assert not r["ok"] and "ゼロ長" in r["error"]

    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"add_connector": {
        "from": {"pic_index": 9}, "to": {"px": [1, 1]}}}]}, output=out)
    assert not r["ok"] and "範囲外" in r["error"]
    assert not (tmp_path / "out.xlsx").exists()


def test_delete_connector_by_index_minimal_touch(tmp_path):
    p = make_book(tmp_path, with_connectors=True)
    out, report = _apply(p, tmp_path, [{"delete_connector": {"index": 2}}])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    assert '<a:stCxn id="101" idx="3"/>' in dx    # 1本目（接続あり）は残る
    assert 'name="コネクタ 111"' not in dx          # 2本目（座標のみ）が消えた
    changed, added = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml"]
    assert added == []

    bad = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"delete_connector": {"index": 9}}]},
                             output=str(tmp_path / "x.xlsx"))
    assert not bad["ok"] and "範囲外" in bad["error"]
    bad = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"delete_connector": {}}]},
                             output=str(tmp_path / "x.xlsx"))
    assert not bad["ok"] and "どちらか一方" in bad["error"]


def test_delete_connector_by_connected_text(tmp_path):
    p = make_book(tmp_path, with_connectors=True)
    out, report = _apply(p, tmp_path, [{"delete_connector": {"connected_to_text": "ラベルA"}}])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    assert "<a:stCxn" not in dx                   # 接続コネクタ 110 が消えた
    assert 'name="コネクタ 111"' in dx              # 座標のみの 111 は無関係なので残る

    bad = xlsm_ops.apply_ops(
        p, {"sheet": "N1", "ops": [{"delete_connector": {"connected_to_text": "重複文言"}}]},
        output=str(tmp_path / "x.xlsx"))
    assert not bad["ok"] and "2 個" in bad["error"]  # 図形テキスト自体が曖昧


def test_delete_connector_multiple_requires_all(tmp_path):
    """同一図形に複数本接続: all なしは明示エラー、all: true で全削除。"""
    p = make_book(tmp_path, with_connectors=True)
    add = {"add_connector": {"from": {"shape_text": "ラベルA"},
                             "to": {"px": [400, 100]}}}

    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [
        add, {"delete_connector": {"connected_to_text": "ラベルA"}}]},
        output=str(tmp_path / "x.xlsx"))
    assert not r["ok"] and "2 本" in r["error"]
    assert not (tmp_path / "x.xlsx").exists()

    out, report = _apply(p, tmp_path, [
        add, {"delete_connector": {"connected_to_text": "ラベルA", "all": True}}])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    assert "<a:stCxn" not in dx and "<a:endCxn" not in dx
    assert 'name="コネクタ 111"' in dx              # 座標のみの 111 は残る


def test_replace_image(tmp_path):
    p = make_book(tmp_path)
    new_png = str(tmp_path / "new.png")
    PIL.new("RGB", (120, 260), (255, 0, 0)).save(new_png)
    out, report = _apply(p, tmp_path, [
        {"replace_image": {"index": 1, "image": new_png}},
    ])
    assert report["ok"], report
    changed, added = _changed(p, out)
    assert changed == ["xl/media/image1.png"]
    assert added == []
    assert entries_of(out)["xl/media/image1.png"] == open(new_png, "rb").read()

    # index 範囲外・非PNG は明示エラー
    bad = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [
        {"replace_image": {"index": 9, "image": new_png}}]}, output=str(tmp_path / "x.xlsx"))
    assert not bad["ok"] and "範囲外" in bad["error"]


def test_clone_sheet_openpyxl_loadable(tmp_path):
    """複製シートが workbook/rels/Content_Types に配線され、Excel 互換で読める。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"clone_sheet": {"template": "template", "new_name": "N2", "tab_after": "N1"}},
        {"set_cell": {"sheet": "N2", "cell": "B3", "text": "複製後の追記"}},
    ], sheet=None)
    assert report["ok"], report
    assert "xl/worksheets/sheet4.xml" in report["added_entries"]

    wb = load_workbook(out)
    assert wb.sheetnames == ["N1", "N2", "改版履歴", "template"]
    assert wb["N2"]["B2"].value == "TPLヘッダ"
    assert wb["N2"]["B3"].value == "複製後の追記"

    # 既存名との衝突は明示エラー
    bad = xlsm_ops.apply_ops(p, {"ops": [
        {"clone_sheet": {"template": "template", "new_name": "N1"}}]},
        output=str(tmp_path / "y.xlsx"))
    assert not bad["ok"] and "既に存在" in bad["error"]


def test_clone_sheet_duplicates_drawing(tmp_path):
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"clone_sheet": {"template": "N1", "new_name": "N1copy"}},
    ], sheet=None)
    assert report["ok"], report
    assert "xl/drawings/drawing2.xml" in report["added_entries"]
    assert "xl/drawings/_rels/drawing2.xml.rels" in report["added_entries"]
    wb = load_workbook(out)
    assert len(wb["N1copy"]._images) == 1  # 複製側にも画像が見える
    assert len(wb["N1"]._images) == 1


def _merges_and_count(sx: str):
    refs = re.findall(r'<mergeCell ref="([^"]+)"/>', sx)
    cm = re.search(r'<mergeCells count="(\d+)">', sx)
    return refs, (int(cm.group(1)) if cm else None)


def test_set_merge_adds_ranges_and_count(tmp_path):
    """ranges 複数指定で結合が張られ、mergeCells count が実体と一致して増える。"""
    from gwex.domains import xlsm_verify

    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"set_merge": {"ranges": ["B14:C17", "D14:F17"]}},
        {"set_merge": {"range": "B26:D26"}},
    ])
    assert report["ok"], report
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    refs, count = _merges_and_count(sx)
    assert {"B14:C17", "D14:F17", "B26:D26"} <= set(refs)
    assert {"B8:D9", "B22:C23"} <= set(refs)      # 既存結合は不変
    assert count == len(refs) == 5                 # 2(既存) + 3(追加)
    assert "解除" not in " ".join(report["ops"])   # 重なり無し → notes なし

    # (d) verify-xlsm の merge 妥当性検査（形式/向き/重複/count 一致）を通る
    ver = xlsm_verify.verify(entries_of(p), entries_of(out),
                             expect_entries=["xl/worksheets/sheet1.xml"])
    assert ver["ok"], ver["issues"]
    assert ver["counts"]["errors"] == 0


def test_set_merge_removes_overlapping_merge_first(tmp_path):
    """既存結合と重なる範囲は先に解除してから張り、解除内容を op 結果に記録する。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"set_merge": {"range": "B8:E10"}},        # 既存 B8:D9 と重なる
    ])
    assert report["ok"], report
    sx = entries_of(out)["xl/worksheets/sheet1.xml"].decode("utf-8")
    refs, count = _merges_and_count(sx)
    assert "B8:E10" in refs and "B8:D9" not in refs
    assert count == len(refs) == 2                 # 2 - 1(解除) + 1(追加)
    assert any("解除" in s and "B8:D9" in s for s in report["ops"]), report["ops"]


def test_set_merge_minimal_touch_and_new_mergecells_element(tmp_path):
    """変更は対象シート XML のみ。mergeCells 要素が無いシートには新設される。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"set_merge": {"sheet": "template", "range": "B2:C3"}},
    ], sheet=None)
    assert report["ok"], report
    changed, added = _changed(p, out)
    assert changed == ["xl/worksheets/sheet3.xml"]  # template シートだけ
    assert added == []
    sx = entries_of(out)["xl/worksheets/sheet3.xml"].decode("utf-8")
    assert '<mergeCells count="1"><mergeCell ref="B2:C3"/></mergeCells>' in sx


def test_set_merge_validation_errors_write_nothing(tmp_path):
    p = make_book(tmp_path)
    out = str(tmp_path / "x.xlsx")
    cases = [
        ({"range": "B8"}, "結合範囲が不正"),           # 単一セル参照（範囲でない）
        ({"range": "B8:B8"}, "1セルだけ"),             # 1セル範囲
        ({"range": "C9:B8"}, "向きが不正"),            # 右下:左上
        ({"range": "B8:C9", "ranges": ["D1:E2"]}, "どちらか一方"),
        ({}, "どちらか一方"),
        ({"ranges": []}, "リスト"),                     # 空リスト
    ]
    for params, msg in cases:
        r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"set_merge": params}]}, output=out)
        assert not r["ok"] and msg in r["error"], (params, r)
    assert not (tmp_path / "x.xlsx").exists()


def test_vba_and_sharedstrings_always_preserved(tmp_path):
    """どの op を通しても vbaProject.bin と sharedStrings はバイト不変。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"set_cell": {"cell": "B8", "text": "x"}},
        {"edit_shape_text": {"match_text": "ラベルA", "new_text": "ラベルB"}},
        {"append_history_row": {"sheet": "改版履歴", "date": "2026-07-11",
                                "target": "N1", "status": "変更", "content": "c"}},
    ])
    assert report["ok"], report
    before, after = entries_of(p), entries_of(out)
    assert after["xl/vbaProject.bin"] == before["xl/vbaProject.bin"]
    assert after["xl/sharedStrings.xml"] == before["xl/sharedStrings.xml"]


def test_unknown_op_and_bad_spec(tmp_path):
    p = make_book(tmp_path, with_image=False, with_shapes=False)
    out = str(tmp_path / "out.xlsx")
    r = xlsm_ops.apply_ops(p, {"ops": [{"nop": {}}]}, output=out)
    assert not r["ok"] and "未知の op" in r["error"]
    r = xlsm_ops.apply_ops(p, {"ops": []}, output=out)
    assert not r["ok"]
    r = xlsm_ops.apply_ops(p, {"ops": [{"set_cell": {"cell": "B8", "text": "x"}}]}, output=out)
    assert not r["ok"] and "sheet が未指定" in r["error"]
    assert not (tmp_path / "out.xlsx").exists()


# フィクスチャ N1 の幾何: cols 指定なし → 列幅 64px=609600EMU、
# 行 ht 指定なし → 既定行高（openpyxl の defaultRowHeight=15pt でも
# フォールバック 20px でも同値）= 190500EMU。
_COL_EMU = 609600
_ROW_EMU = 190500


def test_move_shape_two_cell_anchor_crosses_cell_boundaries(tmp_path):
    """twoCellAnchor のセル跨ぎ移動: from/to とも列幅・行高で col/off を再計算し、
    xfrm off キャッシュは delta 追従・ext は不変。"""
    p = make_book(tmp_path)
    # sp101: from c5+0 r30+0 / to c7+0 r32+0 / xfrm off (1000000, 2000000)
    out, report = _apply(p, tmp_path, [
        {"move_shape": {"match_text": "ボタン変更前", "dx_px": -10, "dy_px": 120}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    # dx=-95250EMU: colOff 0-95250<0 → col-1 して正規化 / dy=+1143000EMU=6行ちょうど
    assert ("<xdr:from><xdr:col>4</xdr:col><xdr:colOff>514350</xdr:colOff>"
            "<xdr:row>36</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>") in dx
    assert ("<xdr:to><xdr:col>6</xdr:col><xdr:colOff>514350</xdr:colOff>"
            "<xdr:row>38</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>") in dx
    assert 514350 == _COL_EMU - 10 * 9525 and 120 * 9525 == 6 * _ROW_EMU
    blk = re.search(r"<xdr:twoCellAnchor>(?:(?!</xdr:twoCellAnchor>).)*ボタン変更前.*?</xdr:twoCellAnchor>",
                    dx, re.DOTALL).group(0)
    assert '<a:off x="904750" y="3143000"/>' in blk        # (1000000-95250, 2000000+1143000)
    assert '<a:ext cx="548640" cy="365760"/>' in blk       # ext は不変
    assert any("move_shape" in s for s in report["ops"])

    changed, added = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml"]
    assert added == []


def test_move_shape_by_id_one_cell_anchor(tmp_path):
    """id:（cNvPr id）直指定で oneCellAnchor の pic を移動（from のみ・ext 不変）。"""
    p = make_book(tmp_path)  # pic capture1: oneCellAnchor from c2+0 r19+0, id=100001
    out, report = _apply(p, tmp_path, [
        {"move_shape": {"id": 100001, "dx_px": 10, "dy_px": 100}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    # 10px=95250EMU は列内に収まる / 100px=952500EMU=5行ちょうど
    assert ("<xdr:from><xdr:col>2</xdr:col><xdr:colOff>95250</xdr:colOff>"
            "<xdr:row>24</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>") in dx
    assert f'<xdr:ext cx="{120 * 9525}" cy="{260 * 9525}"/>' in dx  # 表示サイズ不変
    changed, _ = _changed(p, out)
    assert changed == ["xl/drawings/drawing1.xml"]


def test_move_shape_absolute_anchor(tmp_path):
    """absoluteAnchor は pos を delta EMU 加算し、xfrm off も追従（ext 不変）。"""
    p = make_book(tmp_path)
    out, report = _apply(p, tmp_path, [
        {"add_shape": {"preset": "wedgeRectCallout", "text": "移動対象",
                       "at_px": [120, 340], "size_px": [160, 40],
                       "style_from_text": "ラベルA"}},
        {"move_shape": {"match_text": "移動対象", "dx_px": -20, "dy_px": 60}},
    ])
    assert report["ok"], report
    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    blk = re.search(r"<xdr:absoluteAnchor>.*?</xdr:absoluteAnchor>", dx, re.DOTALL).group(0)
    assert f'<xdr:pos x="{100 * 9525}" y="{400 * 9525}"/>' in blk
    assert f'<xdr:ext cx="{160 * 9525}" cy="{40 * 9525}"/>' in blk  # ext 不変
    assert f'<a:off x="{100 * 9525}" y="{400 * 9525}"/>' in blk     # xfrm キャッシュ追従
    assert f'<a:ext cx="{160 * 9525}" cy="{40 * 9525}"/>' in blk


def test_move_shape_ambiguous_or_bad_params_fail_without_write(tmp_path):
    p = make_book(tmp_path)
    out = str(tmp_path / "out.xlsx")
    cases = [
        ({"match_text": "重複文言", "dy_px": 10}, "2 個"),
        ({"match_text": "存在しない文言", "dy_px": 10}, "一致する図形がありません"),
        ({"id": 9999, "dy_px": 10}, "がありません"),
        ({"match_text": "ラベルA", "id": 102, "dy_px": 10}, "どちらか一方"),
        ({"dy_px": 10}, "どちらか一方"),
        ({"match_text": "ラベルA"}, "移動量が 0"),
        ({"match_text": "ラベルA", "dy_px": "12"}, "整数(px)"),
    ]
    for params, msg in cases:
        r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"move_shape": params}]}, output=out)
        assert not r["ok"] and msg in r["error"], (params, r)
    assert not (tmp_path / "out.xlsx").exists()


def test_move_shape_negative_result_fails_without_write(tmp_path):
    p = make_book(tmp_path)
    out = str(tmp_path / "out.xlsx")
    # 上方向: r30 の原点 30×190500=5715000EMU < 700px=6667500EMU
    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [
        {"move_shape": {"match_text": "ボタン変更前", "dy_px": -700}}]}, output=out)
    assert not r["ok"] and "負座標" in r["error"]
    # 左方向: c5 の原点 5×609600=3048000EMU < 400px=3810000EMU
    r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [
        {"move_shape": {"match_text": "ボタン変更前", "dx_px": -400}}]}, output=out)
    assert not r["ok"] and "負座標" in r["error"]
    assert not (tmp_path / "out.xlsx").exists()


def test_add_image_adds_media_rel_and_pic(tmp_path):
    """新規 media の採番・rels 配線・absoluteAnchor pic の注入。size_px 省略は実サイズ。"""
    from gwex.domains import xlsm_verify

    p = make_book(tmp_path)
    png = str(tmp_path / "modal.png")
    PIL.new("RGB", (32, 20), (10, 20, 30)).save(png)
    out, report = _apply(p, tmp_path, [
        {"add_image": {"image": png, "at_px": [700, 400], "size_px": [64, 40]}},
        {"add_image": {"image": png, "at_px": [800, 400]}},
    ])
    assert report["ok"], report
    changed, added = _changed(p, out)
    assert changed == ["xl/drawings/_rels/drawing1.xml.rels", "xl/drawings/drawing1.xml"]
    assert added == ["xl/media/image2.png", "xl/media/image3.png"]
    assert report["added_entries"] == ["xl/media/image2.png", "xl/media/image3.png"]
    assert entries_of(out)["xl/media/image2.png"] == open(png, "rb").read()

    dx = entries_of(out)["xl/drawings/drawing1.xml"].decode("utf-8")
    b1, b2 = re.findall(r"<xdr:absoluteAnchor>.*?</xdr:absoluteAnchor>", dx, re.DOTALL)
    assert f'<xdr:pos x="{700 * 9525}" y="{400 * 9525}"/>' in b1
    assert f'<xdr:ext cx="{64 * 9525}" cy="{40 * 9525}"/>' in b1
    assert f'<xdr:ext cx="{32 * 9525}" cy="{20 * 9525}"/>' in b2   # 実画像サイズ
    rid1 = re.search(r'r:embed="(rId\d+)"', b1).group(1)
    rels = entries_of(out)["xl/drawings/_rels/drawing1.xml.rels"].decode("utf-8")
    assert re.search(r'Id="%s"[^>]*Target="\.\./media/image2\.png"' % rid1, rels)
    ids = re.findall(r'<xdr:cNvPr [^>]*?id="(\d+)"', dx)
    assert len(ids) == len(set(ids))                               # cNvPr id の一意採番

    ver = xlsm_verify.verify(entries_of(p), entries_of(out), expect_entries=[
        "xl/drawings/drawing1.xml", "xl/drawings/_rels/drawing1.xml.rels",
        "xl/media/image2.png", "xl/media/image3.png"])
    assert ver["ok"], ver["issues"]
    assert ver["counts"]["errors"] == 0


def test_add_image_non_png_or_bad_params_fail_without_write(tmp_path):
    p = make_book(tmp_path)
    jpg = str(tmp_path / "photo.jpg")
    PIL.new("RGB", (10, 10), (1, 2, 3)).save(jpg)
    out = str(tmp_path / "out.xlsx")
    cases = [
        ({"image": jpg, "at_px": [0, 0]}, "PNG ではありません"),
        ({"image": str(tmp_path / "nai.png"), "at_px": [0, 0]}, "がありません"),
        ({"image": jpg, "at_px": [0]}, "2要素"),
        ({"image": jpg, "at_px": [0, 0], "size_px": [10]}, "2要素"),
        ({"image": jpg}, "必須パラメータ"),
    ]
    for params, msg in cases:
        r = xlsm_ops.apply_ops(p, {"sheet": "N1", "ops": [{"add_image": params}]}, output=out)
        assert not r["ok"] and msg in r["error"], (params, r)
    assert not (tmp_path / "out.xlsx").exists()
