"""sheet-scan / apply-ops / verify-xlsm テスト共通のフィクスチャ生成。

openpyxl は「生成時のみ」使用（規約どおり）。図形（xdr:sp）は openpyxl が
書けないため、実ファイル（Excel 由来・xdr: 接頭辞）と同じ形の XML 断片を
ZIP 直接編集で注入する。vbaProject.bin もダミーバイトを注入し、
「VBA を壊さない」ことの検証に使う。
"""

from __future__ import annotations

import re
import zipfile

from openpyxl import Workbook

from gwex.writer import xlsx_zip

IW, IH = 120, 260  # 埋め込みスクショの元ピクセル寸法

_SST = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
    'count="1" uniqueCount="1">'
    '<si><t>■イベント</t><rPh sb="1" eb="5"><t>イベント</t></rPh>'
    '<phoneticPr fontId="1"/></si></sst>'
)
_SST_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
_SST_RTYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"


def _wire_shared_strings(entries: dict[str, bytes]) -> None:
    """B12 を実ファイルと同じ shared string 参照（t="s"・ふりがな rPh つき）にする。

    openpyxl は文字列を inlineStr で書くため、sharedStrings.xml は手で配線する。
    """
    entries["xl/sharedStrings.xml"] = _SST.encode("utf-8")
    ct = entries["[Content_Types].xml"].decode("utf-8")
    ct = ct.replace("</Types>",
                    f'<Override PartName="/xl/sharedStrings.xml" ContentType="{_SST_CT}"/></Types>', 1)
    entries["[Content_Types].xml"] = ct.encode("utf-8")
    xlsx_zip._add_rel(entries, "xl/_rels/workbook.xml.rels", _SST_RTYPE, "sharedStrings.xml")
    sx = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    sx, n = re.subn(r'<c r="B12"[^>]*>.*?</c>', '<c r="B12" t="s"><v>0</v></c>', sx,
                    count=1, flags=re.DOTALL)
    assert n == 1, "B12 セルが見つからない（フィクスチャ生成の前提が崩れた）"
    entries["xl/worksheets/sheet1.xml"] = sx.encode("utf-8")


def entries_of(path: str) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as z:
        return {n: z.read(n) for n in z.namelist()}


def rewrite(path: str, entries: dict[str, bytes]) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n, d in entries.items():
            z.writestr(n, d)


def sp_anchor(id_: int, name: str, text: str, *, prst: str = "wedgeRectCallout",
              row: int = 30, off=(1000000, 2000000), ext=(548640, 365760)) -> str:
    """実ファイルの xdr:sp と同じ構造（twoCellAnchor / xfrm / prstGeom / txBody）。"""
    return (
        f"<xdr:twoCellAnchor><xdr:from><xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f"<xdr:to><xdr:col>7</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{row + 2}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        f'<xdr:sp macro="" textlink=""><xdr:nvSpPr>'
        f'<xdr:cNvPr id="{id_}" name="{name}"/><xdr:cNvSpPr/></xdr:nvSpPr>'
        f'<xdr:spPr><a:xfrm><a:off x="{off[0]}" y="{off[1]}"/>'
        f'<a:ext cx="{ext[0]}" cy="{ext[1]}"/></a:xfrm>'
        f'<a:prstGeom prst="{prst}"><a:avLst/></a:prstGeom>'
        f'<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
        f'<a:ln w="19080"><a:solidFill><a:srgbClr val="FF9999"/></a:solidFill></a:ln></xdr:spPr>'
        f'<xdr:txBody><a:bodyPr anchor="ctr"/><a:lstStyle/>'
        f'<a:p><a:pPr algn="ctr"/><a:r><a:rPr lang="ja-JP" sz="1600" b="1"/>'
        f"<a:t>{text}</a:t></a:r>"
        f'<a:endParaRPr lang="en-US" sz="1600"/></a:p></xdr:txBody></xdr:sp>'
        f"<xdr:clientData/></xdr:twoCellAnchor>"
    )


def cxn_anchor(id_: int, *, prst: str = "straightConnector1",
               st=None, end=None,
               off=(1548640, 2182880), ext=(951360, 0), flips: str = "") -> str:
    """実ファイルの xdr:cxnSp と同じ構造（twoCellAnchor / stCxn・endCxn / 絶対 xfrm /
    w=19050・tx1・tailEnd triangle / style ブロック）。"""
    stx = f'<a:stCxn id="{st[0]}" idx="{st[1]}"/>' if st else ""
    enx = f'<a:endCxn id="{end[0]}" idx="{end[1]}"/>' if end else ""
    return (
        f"<xdr:twoCellAnchor><xdr:from><xdr:col>8</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>30</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f"<xdr:to><xdr:col>9</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>31</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        f'<xdr:cxnSp macro=""><xdr:nvCxnSpPr><xdr:cNvPr id="{id_}" name="コネクタ {id_}"/>'
        f"<xdr:cNvCxnSpPr><a:cxnSpLocks/>{stx}{enx}</xdr:cNvCxnSpPr></xdr:nvCxnSpPr>"
        f'<xdr:spPr><a:xfrm{flips}><a:off x="{off[0]}" y="{off[1]}"/>'
        f'<a:ext cx="{ext[0]}" cy="{ext[1]}"/></a:xfrm>'
        f'<a:prstGeom prst="{prst}"><a:avLst/></a:prstGeom>'
        f'<a:ln w="19050"><a:solidFill><a:schemeClr val="tx1"/></a:solidFill>'
        f'<a:tailEnd type="triangle"/></a:ln></xdr:spPr>'
        f'<xdr:style><a:lnRef idx="1"><a:schemeClr val="accent1"/></a:lnRef>'
        f'<a:fillRef idx="0"><a:schemeClr val="accent1"/></a:fillRef>'
        f'<a:effectRef idx="0"><a:schemeClr val="accent1"/></a:effectRef>'
        f'<a:fontRef idx="minor"><a:schemeClr val="tx1"/></a:fontRef></xdr:style>'
        f"</xdr:cxnSp><xdr:clientData/></xdr:twoCellAnchor>"
    )


def make_book(tmp_path, *, with_image: bool = True, with_shapes: bool = True,
              with_connectors: bool = False) -> str:
    """N1（画面シート）＋改版履歴＋template の3シート構成のフィクスチャ。

    - N1: B列 ■マーカー2つ / 結合セル B8:D9・B22:C23 / 非B非結合の C3
    - 改版履歴: row4 ヘッダ・row5 最終データ行（A=シリアル値）
    - drawing1.xml: pic 1枚（capture1, oneCellAnchor row0=19）+ sp 図形
    - with_connectors: 実ファイル形状の cxnSp 2本（110=101→102 接続 / 111=座標のみ）と、
      pic への a:xfrm 絶対座標キャッシュ（Excel 保存済み実ファイルは必ず持つ）を注入
    - xl/vbaProject.bin: ダミーバイト
    """
    import PIL.Image

    wb = Workbook()
    ws = wb.active
    ws.title = "N1"
    ws["B7"] = "■画面概要"
    ws["B8"] = "ダッシュボード概要"
    ws.merge_cells("B8:D9")
    ws["C3"] = "タイトル"
    ws["A4"] = 46000
    ws["B12"] = "■イベント"
    ws["B22"] = "下部データ"
    ws.merge_cells("B22:C23")

    kaihan = wb.create_sheet("改版履歴")
    kaihan["A4"] = 43040
    kaihan["B4"] = "対象"
    kaihan["C4"] = "状態"
    kaihan["E4"] = "内容"
    kaihan["A5"] = 46171
    kaihan["B5"] = "N0501"
    kaihan["C5"] = "変更"
    kaihan["E5"] = "旧内容"

    tpl = wb.create_sheet("template")
    tpl["B2"] = "TPLヘッダ"

    p = str(tmp_path / "base.xlsx")
    wb.save(p)

    if with_image:
        shot = str(tmp_path / "shot.png")
        PIL.Image.new("RGB", (IW, IH), (0, 128, 255)).save(shot)
        xlsx_zip.set_cell_image(p, "N1", "C20", shot, width=IW, height=IH)

    entries = entries_of(p)
    _wire_shared_strings(entries)
    if with_shapes:
        dx = entries["xl/drawings/drawing1.xml"].decode("utf-8")
        # 101 と 102 は接続コネクタの幾何検証用に水平に離して置く
        # （101 右辺中点 (1548640, 2182880) → 102 左辺中点 (2500000, 2182880)）
        shapes = (
            sp_anchor(101, "CustomShape 1", "ボタン変更前", off=(1000000, 2000000))
            + sp_anchor(102, "CustomShape 2", "ラベルA", prst="accentBorderCallout2",
                        row=34, off=(2500000, 2000000))
            + sp_anchor(103, "CustomShape 3", "重複文言", row=38)
            + sp_anchor(104, "CustomShape 4", "重複文言", row=42)
        )
        if with_connectors:
            shapes += (
                cxn_anchor(110, st=(101, 3), end=(102, 1))
                + cxn_anchor(111, prst="bentConnector3",
                             off=(1000000, 3000000), ext=(400000, 300000))
            )
        entries["xl/drawings/drawing1.xml"] = dx.replace(
            "</xdr:wsDr>", shapes + "</xdr:wsDr>").encode("utf-8")
    if with_connectors and with_image:
        # Excel 保存済み実ファイルの pic は絶対座標キャッシュ a:xfrm を必ず持つ。
        # gwex の set_cell_image は書かないため、実ファイル形状に寄せて注入する。
        dx = entries["xl/drawings/drawing1.xml"].decode("utf-8")
        a_ns = 'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        xfrm = (f'<a:xfrm {a_ns}><a:off x="1200000" y="8000000"/>'
                f'<a:ext cx="{IW * 9525}" cy="{IH * 9525}"/></a:xfrm>')
        dx = dx.replace("<xdr:spPr><a:prstGeom", f"<xdr:spPr>{xfrm}<a:prstGeom", 1)
        entries["xl/drawings/drawing1.xml"] = dx.encode("utf-8")
    entries["xl/vbaProject.bin"] = b"\x01\x02FAKE-VBA-BYTES\x03"
    rewrite(p, entries)
    return p
