"""xlsx_zip.set_cell_image: 既存の埋め込み画像を壊さず over-cell 画像を注入することの回帰。

openpyxl の load+save は既存 media を落とすため、実物の設計書（既存スクショ多数）には
使えない。本テストは「既存画像を保持」「枠フィット結合」「insert_rows シフト」「base 経由」
を検証する。
"""

from __future__ import annotations

import zipfile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage

from gwex.writer import base, xlsx_zip

PIL = pytest.importorskip("PIL.Image")


def _png(path, size, color):
    PIL.new("RGB", size, color).save(path)
    return str(path)


def _media(path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        return sorted(
            (n for n in z.namelist() if n.startswith("xl/media/")),
            key=lambda x: (len(x), x),
        )


def _base_with_image(tmp_path):
    """既存画像1枚＋テキスト＋結合を持つ xlsx を作る。"""
    red = _png(tmp_path / "red.png", (40, 40), (255, 0, 0))
    p = tmp_path / "base.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "header"
    ws["B10"] = "keep-me"
    ws.merge_cells("A20:B20")
    ws["A20"] = "merged20"
    img = XLImage(red)
    img.anchor = "E2"
    ws.add_image(img)
    wb.save(p)
    return str(p)


def test_existing_image_preserved(tmp_path):
    base_xlsx = _base_with_image(tmp_path)
    shot = _png(tmp_path / "shot.png", (300, 650), (0, 128, 255))
    assert len(_media(base_xlsx)) == 1

    out = str(tmp_path / "out.xlsx")
    xlsx_zip.set_cell_image(base_xlsx, "S1", "H2", shot, max_dim=200, output=out)

    # 既存 image1 を保持しつつ image2 を追加
    media = _media(out)
    assert len(media) == 2
    with zipfile.ZipFile(base_xlsx) as zb, zipfile.ZipFile(out) as zo:
        assert zo.read("xl/media/image1.png") == zb.read("xl/media/image1.png")
    # openpyxl からも 2 枚見える＝drawing に両アンカーがある
    assert len(load_workbook(out)["S1"]._images) == 2


def test_cell_range_merges_and_fits(tmp_path):
    base_xlsx = _base_with_image(tmp_path)
    shot = _png(tmp_path / "shot.png", (300, 650), (0, 128, 255))
    out = str(tmp_path / "out.xlsx")
    xlsx_zip.set_cell_image(base_xlsx, "S1", "C5", shot, cell_range="C5:F25", max_dim=1024, output=out)

    ws = load_workbook(out)["S1"]
    assert "C5:F25" in [str(m) for m in ws.merged_cells.ranges]
    assert len(ws._images) == 2  # 既存＋新規


def test_insert_rows_shifts_existing_content(tmp_path):
    base_xlsx = _base_with_image(tmp_path)
    shot = _png(tmp_path / "shot.png", (300, 650), (0, 128, 255))
    out = str(tmp_path / "out.xlsx")
    # C5:F25 = 21 行。min_r=5 なので 5 以降が +21 される
    xlsx_zip.set_cell_image(base_xlsx, "S1", "C5", shot, cell_range="C5:F25",
                            insert_rows=True, max_dim=1024, output=out)

    ws = load_workbook(out)["S1"]
    assert ws["A1"].value == "header"          # min_r 未満は不動
    assert ws["B31"].value == "keep-me"        # B10 -> B31
    assert ws["A41"].value == "merged20"       # A20 -> A41
    assert "A41:B41" in [str(m) for m in ws.merged_cells.ranges]
    assert "C5:F25" in [str(m) for m in ws.merged_cells.ranges]


def test_base_routes_local_xlsx_to_zip(tmp_path):
    """base.set_cell_image（ローカル .xlsx）は既存画像を保持する＝xlsx_zip 経由。"""
    base_xlsx = _base_with_image(tmp_path)
    shot = _png(tmp_path / "shot.png", (300, 650), (0, 128, 255))
    out = str(tmp_path / "out.xlsx")
    base.set_cell_image(base_xlsx, "S1", "C5", shot, cell_range="C5:F25", max_dim=512, output=out)

    assert len(_media(out)) == 2  # 既存画像が残っている（openpyxl 経由なら 1 に落ちる）
    assert len(load_workbook(out)["S1"]._images) == 2


def test_anchor_position_without_range(tmp_path):
    """cell_range 無しのアンカー指定が (col,row) を取り違えず正しい位置に置かれること。

    （coordinate_to_tuple は (row, col) を返す。逆代入すると列=行番号になり画面外へ飛ぶ回帰。）
    """
    shot = _png(tmp_path / "shot.png", (120, 260), (0, 128, 255))
    p = tmp_path / "base.xlsx"
    wb = Workbook()
    wb.active.title = "S1"
    wb.save(p)
    out = str(tmp_path / "out.xlsx")
    xlsx_zip.set_cell_image(str(p), "S1", "C210", shot, output=out)

    ws = load_workbook(out)["S1"]
    assert len(ws._images) == 1
    frm = ws._images[0].anchor._from  # oneCellAnchor の from
    assert (frm.col, frm.row) == (2, 209)  # C=col2(0始まり), 210行=row209


def test_idempotent_two_injections_accumulate(tmp_path):
    """連続注入で画像が増えていく（前段の出力を次段が読む）。"""
    base_xlsx = _base_with_image(tmp_path)
    shot = _png(tmp_path / "shot.png", (300, 650), (0, 128, 255))
    s1 = str(tmp_path / "s1.xlsx")
    s2 = str(tmp_path / "s2.xlsx")
    xlsx_zip.set_cell_image(base_xlsx, "S1", "C5", shot, max_dim=300, output=s1)
    xlsx_zip.set_cell_image(s1, "S1", "H5", shot, max_dim=300, output=s2)
    assert len(_media(s2)) == 3
    assert len(load_workbook(s2)["S1"]._images) == 3
