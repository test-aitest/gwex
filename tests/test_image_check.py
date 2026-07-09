"""image_check（画像のはみ出し/浮遊/二重/前後サイズ違いの検出）の検証。

概要設計書「2.画面イメージ」で実際に起きた破綻を人工的に再現し、
目視でなく数値で検出できることを確認する。
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XLImage

from gwex.domains import image_check

PIL = pytest.importorskip("PIL.Image")


def _png(size) -> bytes:
    buf = BytesIO()
    PIL.new("RGB", size, (255, 0, 0)).save(buf, format="PNG")
    return buf.getvalue()


def _add(ws, anchor: str, w: int, h: int, src: tuple[int, int] | None = None) -> None:
    """表示寸法 w x h px の画像を anchor に貼る。

    src（元画像の寸法）を省略すると表示寸法と同比にする＝縦横比を保った配置。
    aspect 検査を意図的に発火させたいときだけ src を明示する。
    """
    img = XLImage(BytesIO(_png(src or (w, h))))
    img.anchor = anchor
    img.width, img.height = w, h
    ws.add_image(img)


def _book(tmp_path, build) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2.画面イメージ(iOS)"
    build(ws)
    p = tmp_path / "t.xlsx"
    wb.save(p)
    return str(p)


def _kinds(path):
    return sorted(i["kind"] for i in image_check.check(path)["issues"])


def test_ok_pair_in_frames(tmp_path):
    """修正前/修正後が同寸で、それぞれの結合枠に収まっていれば破綻ゼロ。"""
    def build(ws):
        ws.merge_cells("C8:G20")
        ws.merge_cells("H8:L20")
        _add(ws, "C8", 100, 120)
        _add(ws, "H8", 100, 120)
    path = _book(tmp_path, build)
    result = image_check.check(path)
    assert result["issues"] == []
    assert len(result["images"]["2.画面イメージ(iOS)"]) == 2


def test_overflow_detected(tmp_path):
    """枠より明らかに大きい画像は overflow として検出される。"""
    def build(ws):
        ws.merge_cells("C8:D9")  # 小さな枠
        _add(ws, "C8", 900, 900)
    assert "overflow" in _kinds(_book(tmp_path, build))


def test_pair_mismatch_detected(tmp_path):
    """同じ行の修正前/修正後で表示寸法が違えば pair_mismatch。"""
    def build(ws):
        ws.merge_cells("C8:G20")
        ws.merge_cells("H8:L20")
        _add(ws, "C8", 100, 120)
        _add(ws, "H8", 140, 120)   # 幅だけ違う
    assert "pair_mismatch" in _kinds(_book(tmp_path, build))


def test_two_pairs_in_one_row_are_compared_pairwise(tmp_path):
    """1行に左ペア(C/H)と右ペア(N/S)が並ぶテンプレ構造。

    ペアごとに前後が揃っていれば、左右でサイズが違っても破綻ではない
    （実物の 2026-00000x1 がこの形。行単位で比較すると誤検出になる）。
    """
    def build(ws):
        for a, b in [("C8", "G20"), ("H8", "L20"), ("N8", "R20"), ("S8", "W20")]:
            ws.merge_cells(f"{a}:{b}")
        _add(ws, "C8", 120, 200)   # 左ペア
        _add(ws, "H8", 120, 200)
        _add(ws, "N8", 60, 200)    # 右ペア（左と幅が違うが、ペア内では一致）
        _add(ws, "S8", 60, 200)
    assert image_check.check(_book(tmp_path, build))["issues"] == []


def test_pair_mismatch_within_right_pair(tmp_path):
    """右ペアの中で前後サイズが違えば、左ペアが揃っていても検出する。"""
    def build(ws):
        for a, b in [("C8", "G20"), ("H8", "L20"), ("N8", "R20"), ("S8", "W20")]:
            ws.merge_cells(f"{a}:{b}")
        _add(ws, "C8", 120, 200)
        _add(ws, "H8", 120, 200)
        _add(ws, "N8", 60, 200)
        _add(ws, "S8", 90, 200)    # 右ペアの後だけ違う
    issues = image_check.check(_book(tmp_path, build))["issues"]
    assert [i["kind"] for i in issues] == ["pair_mismatch"]
    assert issues[0]["cell"] == "N8"


def test_subpixel_pair_diff_is_tolerated(tmp_path):
    """set-image の丸め誤差（1px 前後）は破綻としない。

    実データ由来: 2026-000999 の answer（ユーザー提供の正解）が
    C8=306.8x669.0px / H8=306.2x667.7px で 0.6/1.3px ずれている。
    """
    def build(ws):
        ws.merge_cells("C8:G45")
        ws.merge_cells("H8:L45")
        _add(ws, "C8", 307, 669)
        _add(ws, "H8", 306, 668)
    assert image_check.check(_book(tmp_path, build))["issues"] == []


def test_visible_pair_diff_is_flagged(tmp_path):
    """許容を超える差（実データ 2026-000322 の高さ 4px 差）は検出する。"""
    def build(ws):
        ws.merge_cells("C8:G45")
        ws.merge_cells("H8:L45")
        _add(ws, "C8", 306, 667)
        _add(ws, "H8", 304, 663)   # 高さ 4px 差
    assert "pair_mismatch" in _kinds(_book(tmp_path, build))


def test_duplicate_detected(tmp_path):
    """同一アンカーに2枚（埋め込み＋overlay の二重貼り）は duplicate。"""
    def build(ws):
        ws.merge_cells("C8:G20")
        _add(ws, "C8", 100, 120)
        _add(ws, "C8", 100, 120)
    assert "duplicate" in _kinds(_book(tmp_path, build))


def test_no_frame_detected(tmp_path):
    """結合枠の左上に紐づかない画像は no_frame（セル上に浮いている）。"""
    def build(ws):
        _add(ws, "C8", 100, 120)   # merge していない
    assert "no_frame" in _kinds(_book(tmp_path, build))


def test_aspect_distortion_detected(tmp_path):
    """元画像 40x30 を縦横比を無視して引き伸ばすと aspect として検出される。"""
    def build(ws):
        ws.merge_cells("C8:G20")
        _add(ws, "C8", 200, 40, src=(40, 30))   # 元比 1.333 に対し表示比 5.0
    assert "aspect" in _kinds(_book(tmp_path, build))


def test_aspect_preserved_not_flagged(tmp_path):
    """縦横比を保った拡大は aspect を出さない（元 40x30 → 表示 120x90）。"""
    def build(ws):
        ws.merge_cells("C8:G20")
        _add(ws, "C8", 120, 90, src=(40, 30))
    assert "aspect" not in _kinds(_book(tmp_path, build))


def test_tolerance_absorbs_small_overflow(tmp_path):
    """px 換算誤差レベルのはみ出しは tolerance_px で吸収され検出しない。"""
    def build(ws):
        ws.merge_cells("C8:C8")
        _add(ws, "C8", 1, 1)
    path = _book(tmp_path, build)
    assert image_check.check(path, tolerance_px=100.0)["issues"] == []
