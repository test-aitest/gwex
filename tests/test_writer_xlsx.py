"""Excel 書き戻し（set_cell_text / write_cells / last_data_row）のテスト。

実ファイルは触らず、tmp に作った xlsx に対して書き込み→再オープン検証する。
（画像注入は writer/xlsx_zip に一本化したため tests/test_xlsx_zip_image.py を参照）
"""

from __future__ import annotations

import openpyxl
import pytest

from gwex.writer import xlsx_writer


@pytest.fixture
def book(tmp_path):
    p = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "S"
    wb.save(p)
    return str(p)


def test_set_cell_text_inplace(book):
    xlsx_writer.set_cell_text(book, "S", "B2", "こんにちは")
    wb = openpyxl.load_workbook(book)
    assert wb["S"]["B2"].value == "こんにちは"


def test_set_cell_text_output_copy(book, tmp_path):
    out = str(tmp_path / "copy.xlsx")
    xlsx_writer.set_cell_text(book, "S", "A1", "x", output=out)
    # 元ファイルは変わらず、コピーにのみ書かれる
    assert openpyxl.load_workbook(book)["S"]["A1"].value is None
    assert openpyxl.load_workbook(out)["S"]["A1"].value == "x"


def test_write_cells(book):
    xlsx_writer.write_cells(book, "S", [("A1", "名前"), ("B1", 1), ("A2", "値")])
    wb = openpyxl.load_workbook(book)
    assert wb["S"]["A1"].value == "名前"
    assert wb["S"]["B1"].value == 1


