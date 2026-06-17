"""xlsx parser の画像読み取り（埋め込み画像 → Image + Document.media）の検証。

メモリ上で「テキスト＋埋め込み画像」を持つ .xlsx を生成し、parse が
Image ブロックと base64 の MediaItem を返すことを確認する。
"""

from __future__ import annotations

import base64
from io import BytesIO

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XLImage

from gwex.ir import FetchedRef, Image
from gwex.parsers import xlsx

PIL = pytest.importorskip("PIL.Image")


def _png_bytes(size=(40, 30), color=(255, 0, 0)) -> bytes:
    buf = BytesIO()
    PIL.new("RGB", size, color).save(buf, format="PNG")
    return buf.getvalue()


def _xlsx_with_image(png: bytes, *, title: str = "S1", anchor: str = "E2") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws["A1"] = "項目"
    ws["B1"] = "値"
    ws["A2"] = "売上"
    ws["B2"] = 100
    img = XLImage(BytesIO(png))
    img.anchor = anchor
    ws.add_image(img)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def test_embedded_image_extracted_to_media():
    png = _png_bytes()
    doc = xlsx.parse(_xlsx_with_image(png))

    # 本文に Image(FetchedRef) ブロックがある
    images = [b for b in doc.blocks if isinstance(b, Image)]
    assert len(images) == 1
    ref = images[0].src
    assert isinstance(ref, FetchedRef)
    assert ref.mime == "image/png"

    # media に1件、base64 を decode すると PNG マジックで始まる
    assert len(doc.media) == 1
    item = doc.media[0]
    assert item.id == ref.id
    assert item.mime == "image/png"
    raw = base64.b64decode(item.data_b64)
    assert raw[:8] == b"\x89PNG\r\n\x1a\n"


def test_same_image_twice_is_deduped():
    png = _png_bytes()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "x"
    for anchor in ("E2", "E20"):
        img = XLImage(BytesIO(png))
        img.anchor = anchor
        ws.add_image(img)
    out = BytesIO()
    wb.save(out)

    doc = xlsx.parse(out.getvalue())
    images = [b for b in doc.blocks if isinstance(b, Image)]
    # 2 箇所に同一バイト → Image は2つだが media は1件（同 id を共有）
    assert len(images) == 2
    assert len(doc.media) == 1
    assert {img.src.id for img in images} == {doc.media[0].id}


def test_no_image_keeps_media_empty():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S1"
    ws.append(["a", "b"])
    out = BytesIO()
    wb.save(out)

    doc = xlsx.parse(out.getvalue())
    assert doc.media == []
    assert not any(isinstance(b, Image) for b in doc.blocks)
