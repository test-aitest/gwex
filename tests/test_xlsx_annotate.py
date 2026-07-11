"""xlsx_zip.add_annotations / remove_annotations: 画像上の赤枠+番号図形の注入と除去。

openpyxl は図形追加 API を持たず load→save で図形を落とすため、ZIP 直接編集で
drawing XML に xdr:sp を追記する。本テストは「EMU 換算の厳密性」「最小タッチ
（drawing 以外はバイト同一）」「openpyxl / check-images 互換」「可逆性」を検証する。
"""

from __future__ import annotations

import re
import zipfile
from xml.dom import minidom

import pytest
from openpyxl import Workbook, load_workbook

from gwex.domains import image_check
from gwex.writer import xlsx_zip

PIL = pytest.importorskip("PIL.Image")

IW, IH = 300, 650  # 元画像のピクセル寸法


def _png(path, size=(IW, IH), color=(0, 128, 255)):
    PIL.new("RGB", size, color).save(path)
    return str(path)


def _base_with_capture(tmp_path, n_images=1):
    """xlsx_zip.set_cell_image で captureN を注入した xlsx（実ファイルと同じ生成経路）。"""
    shot = _png(tmp_path / "shot.png")
    p = str(tmp_path / "base.xlsx")
    wb = Workbook()
    wb.active.title = "S1"
    wb.save(p)
    anchors = ["C8", "H8"]
    for i in range(n_images):
        xlsx_zip.set_cell_image(p, "S1", anchors[i], shot, width=IW, height=IH)
    return p


def _drawing_xml(path) -> str:
    with zipfile.ZipFile(path) as z:
        names = [n for n in z.namelist() if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n)]
        assert len(names) == 1
        return z.read(names[0]).decode("utf-8")


def _entries(path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as z:
        return {n: z.read(n) for n in z.namelist()}


def test_add_annotation_emu_exact(tmp_path):
    """注入された sp のオフセット・寸法が「表示EMU×画像内比率」の厳密値と一致する。"""
    p = _base_with_capture(tmp_path)
    pic = xlsx_zip._pics_in_drawing(_drawing_xml(p))[0]
    x, y, w, h = 30, 130, 120, 60

    out = str(tmp_path / "out.xlsx")
    xlsx_zip.add_annotations(p, "S1", [(x, y, w, h, "1")], output=out)

    dx = _drawing_xml(out)
    doc = minidom.parseString(dx)  # well-formed であること
    assert doc.documentElement is not None

    x_ = r"(?:xdr:)?"
    rect = re.search(
        rf'<{x_}absoluteAnchor><{x_}pos x="(\d+)" y="(\d+)"/><{x_}ext cx="(\d+)" cy="(\d+)"/>'
        rf'<{x_}sp><{x_}nvSpPr><{x_}cNvPr id="\d+" name="annot1"/>', dx)
    assert rect, "annot1 の sp が注入されていない"
    px, py, cx, cy = (int(g) for g in rect.groups())
    # 絶対座標 = セル原点（Excel 実式換算）+ 画像オフセット + 画像内比率換算
    with zipfile.ZipFile(out) as z:
        sx = z.read("xl/worksheets/sheet1.xml").decode()
    ox, oy = xlsx_zip._anchor_origin_emu(sx, pic["col"], pic["row"])
    assert px == ox + pic["col_off"] + x * pic["cx"] // IW
    assert py == oy + pic["row_off"] + y * pic["cy"] // IH
    assert cx == w * pic["cx"] // IW
    assert cy == h * pic["cy"] // IH

    # 赤枠: noFill + 赤線、バッジ: ellipse + 白番号
    assert '<a:noFill xmlns:a=' in dx
    assert re.search(r'<a:ln [^>]*w="28575"><a:solidFill><a:srgbClr val="FF0000"/>', dx)
    assert 'name="annot1badge"' in dx
    assert 'prst="ellipse"' in dx
    assert "<a:t>1</a:t>" in dx
    # 簡易レンダラ（Quick Look 等）向けの絶対座標 xfrm を持つ
    assert re.search(r'<a:xfrm [^>]*><a:off x="\d+" y="\d+"/><a:ext cx="%d" cy="%d"/></a:xfrm>'
                     % (cx, cy), dx)


def test_minimal_touch_and_compat(tmp_path):
    """drawing 以外の ZIP エントリはバイト同一。openpyxl / check-images も無影響。"""
    p = _base_with_capture(tmp_path)
    out = str(tmp_path / "out.xlsx")
    xlsx_zip.add_annotations(p, "S1", [(10, 10, 50, 40, "1"), (20, 300, 80, 60, "2")], output=out)

    before, after = _entries(p), _entries(out)
    assert set(before) == set(after)
    changed = [n for n in before if before[n] != after[n]]
    assert changed == ["xl/drawings/drawing1.xml"]

    ws = load_workbook(out)["S1"]
    assert len(ws._images) == 1  # 図形は画像として数えられない・load も壊れない

    # check-images の結果が注釈の前後で変わらない（図形は検査対象外）
    assert image_check.check(p, sheets=["S1"]) == image_check.check(out, sheets=["S1"])


def test_labels_and_multiple_rects(tmp_path):
    """複数 --rect の連番と、追記実行時の連番継続。"""
    p = _base_with_capture(tmp_path)
    xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1"), (0, 20, 10, 10, "2")])
    xlsx_zip.add_annotations(p, "S1", [(0, 40, 10, 10, "3")])  # 追記

    dx = _drawing_xml(p)
    assert [m for m in re.findall(r'name="(annot\d+)"', dx)] == ["annot1", "annot2", "annot3"]
    # cNvPr id が重複しない
    ids = re.findall(r'<(?:xdr:)?cNvPr [^>]*id="(\d+)"', dx)
    assert len(ids) == len(set(ids))


def test_name_selection(tmp_path):
    """複数画像で name 未指定はエラー（候補提示）、指定時は対象画像基準で配置。"""
    p = _base_with_capture(tmp_path, n_images=2)
    with pytest.raises(ValueError, match="capture1.*capture2"):
        xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1")])
    with pytest.raises(ValueError, match="見つかりません"):
        xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1")], name="capture9")

    xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1")], name="capture2")
    dx = _drawing_xml(p)
    pic2 = [q for q in xlsx_zip._pics_in_drawing(dx) if q["name"] == "capture2"][0]
    rect = re.search(r'<(?:xdr:)?pos x="(\d+)" y="(\d+)"/>', dx)
    with zipfile.ZipFile(p) as z:
        sx = z.read("xl/worksheets/sheet1.xml").decode()
    ox, _ = xlsx_zip._anchor_origin_emu(sx, pic2["col"], pic2["row"])
    assert int(rect.group(1)) == ox + pic2["col_off"]  # capture2（H8）側の絶対座標


def test_remove_annotations_reversible(tmp_path):
    """remove で sp が全て消え、pic と他エントリは元のまま（可逆性）。"""
    p = _base_with_capture(tmp_path)
    before = _entries(p)
    xlsx_zip.add_annotations(p, "S1", [(10, 10, 50, 40, "1")])

    dest, n = xlsx_zip.remove_annotations(p, "S1")
    assert n == 2  # rect + badge
    after = _entries(dest)
    assert "sp>" not in _drawing_xml(dest)
    assert re.search(r"<(?:xdr:)?pic>", _drawing_xml(dest))  # 画像は残る
    assert before.keys() == after.keys()
    assert all(before[k] == after[k] for k in before if "drawing" not in k)


def test_no_badge_and_color(tmp_path):
    p = _base_with_capture(tmp_path)
    xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1")], badge=False, color="#00b050")
    dx = _drawing_xml(p)
    assert "badge" not in dx
    assert 'val="00B050"' in dx
    with pytest.raises(ValueError, match="RRGGBB"):
        xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1")], color="red")


def test_uses_pic_xfrm_when_present(tmp_path):
    """pic に a:xfrm（Excel の絶対座標キャッシュ）があれば、そこから誤差ゼロで計算する。"""
    p = _base_with_capture(tmp_path)
    with zipfile.ZipFile(p) as z:
        entries = {n: z.read(n) for n in z.namelist()}
    dx0 = entries["xl/drawings/drawing1.xml"].decode()
    dx0 = dx0.replace(
        "<xdr:spPr><a:prstGeom",
        '<xdr:spPr><a:xfrm><a:off x="1000000" y="2000000"/><a:ext cx="1" cy="1"/></a:xfrm>'
        "<a:prstGeom", 1)
    entries["xl/drawings/drawing1.xml"] = dx0.encode()
    with zipfile.ZipFile(p, "w", zipfile.ZIP_DEFLATED) as z:
        for n, d in entries.items():
            z.writestr(n, d)

    pic = xlsx_zip._pics_in_drawing(_drawing_xml(p))[0]
    assert (pic["off_x"], pic["off_y"]) == (1000000, 2000000)
    xlsx_zip.add_annotations(p, "S1", [(10, 20, 30, 40, "1")])
    dx = _drawing_xml(p)
    m = re.search(r'<(?:xdr:)?pos x="(\d+)" y="(\d+)"/>', dx)
    assert int(m.group(1)) == 1000000 + 10 * pic["cx"] // IW
    assert int(m.group(2)) == 2000000 + 20 * pic["cy"] // IH


def test_no_image_sheet_raises(tmp_path):
    p = str(tmp_path / "empty.xlsx")
    wb = Workbook()
    wb.active.title = "S1"
    wb.save(p)
    with pytest.raises(ValueError, match="画像がありません"):
        xlsx_zip.add_annotations(p, "S1", [(0, 0, 10, 10, "1")])
    dest, n = xlsx_zip.remove_annotations(p, "S1")
    assert n == 0
