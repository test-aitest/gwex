"""testspec 抽出（merge_fill + マッピング → 階層 TestSpec）の単体テスト。"""

from __future__ import annotations

from io import BytesIO

import openpyxl

from gwmd.domains import testspec


def _build() -> bytes:
    """既定マッピング（header_row=7, data=10, C/E/G/H/I/J）に沿う合成シート。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3.結合テスト項目"
    ws["C7"], ws["E7"], ws["G7"] = "画面名", "中項目名", "小項目名"
    ws["H7"], ws["I7"], ws["J7"] = "No", "確認内容", "実施手順"
    # 画面A / 中項目X — 確認内容2件（C・E・J は縦結合で1回だけ）
    ws["C10"], ws["E10"] = "画面A", "中項目X"
    ws["H10"], ws["I10"] = 1, "確認1"
    ws["J10"] = "①手順1\n②手順2"
    ws["H11"], ws["I11"] = 2, "確認2"
    ws.merge_cells("C10:C11")
    ws.merge_cells("E10:E11")
    ws.merge_cells("J10:J11")
    # 画面B / 中項目Y — 1件
    ws["C12"], ws["E12"] = "画面B", "中項目Y"
    ws["H12"], ws["I12"] = 1, "確認3"
    ws["J12"] = "単一手順"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_hierarchy_and_merge_fill():
    spec = testspec.extract(_build(), "3.結合テスト項目")
    assert spec.sheet == "3.結合テスト項目"
    assert [s.screen_name for s in spec.screens] == ["画面A", "画面B"]

    a = spec.screens[0]
    assert len(a.groups) == 1
    g = a.groups[0]
    assert g.medium_category == "中項目X"
    assert [c.test_no for c in g.cases] == [1, 2]
    # 縦結合された実施手順が両ケースに前方補完され、改行で配列化
    assert g.cases[0].execution_steps == ["①手順1", "②手順2"]
    assert g.cases[1].execution_steps == ["①手順1", "②手順2"]
    assert g.cases[0].verification_content == "確認1"

    b = spec.screens[1]
    assert b.groups[0].medium_category == "中項目Y"
    assert b.groups[0].cases[0].execution_steps == ["単一手順"]


def test_marker_split_when_no_newline():
    # 改行が無く ①② が連続する場合はマーカーで分割
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["C7"], ws["E7"], ws["G7"] = "画面名", "中項目名", "小項目名"
    ws["H7"], ws["I7"], ws["J7"] = "No", "確認内容", "実施手順"
    ws["C10"], ws["E10"], ws["H10"], ws["I10"] = "画面", "中", 1, "確認"
    ws["J10"] = "①一つ目②二つ目"
    buf = BytesIO()
    wb.save(buf)
    spec = testspec.extract(buf.getvalue(), "S")
    assert spec.screens[0].groups[0].cases[0].execution_steps == ["①一つ目", "②二つ目"]
